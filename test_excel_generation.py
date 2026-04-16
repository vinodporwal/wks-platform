"""
Test script to verify Excel generation is working correctly
"""
import os
import sys

# Add the PPPython-script directory to path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'apps', 'python', 'PPPython-script'))

from openpyxl import load_workbook

# Path to the generated Excel file from the latest run
# Update this path based on your actual log output
excel_path = r"D:\Honeywell\Scripts\Scripts\Python\PPPython-script\PPPython-script\PPPython-script\logs\full_year_run\run_20260324_200754\Annual_Balance_Summary_FY2026_2027.xlsx"

# Check if file exists
if not os.path.exists(excel_path):
    print(f"ERROR: Excel file not found at: {excel_path}")
    print("\nPlease update the excel_path variable with the correct path to your generated Excel file")
    sys.exit(1)

# Load the workbook
print(f"Loading Excel file: {excel_path}")
wb = load_workbook(excel_path)

print(f"\n{'='*70}")
print(f"EXCEL FILE ANALYSIS")
print(f"{'='*70}")
print(f"File size: {os.path.getsize(excel_path):,} bytes ({os.path.getsize(excel_path)/1024:.2f} KB)")
print(f"Number of sheets: {len(wb.sheetnames)}")
print(f"\nSheet names:")
for i, sheet_name in enumerate(wb.sheetnames, 1):
    ws = wb[sheet_name]
    # Count non-empty cells
    non_empty_cells = sum(1 for row in ws.iter_rows() for cell in row if cell.value is not None)
    print(f"  {i}. {sheet_name} - {ws.max_row} rows, {ws.max_column} cols, {non_empty_cells} non-empty cells")

# Check first sheet in detail
if wb.sheetnames:
    first_sheet = wb[wb.sheetnames[0]]
    print(f"\nFirst sheet '{wb.sheetnames[0]}' preview (first 10 rows):")
    for i, row in enumerate(first_sheet.iter_rows(max_row=10, values_only=True), 1):
        print(f"  Row {i}: {row}")

wb.close()
print(f"\n{'='*70}")
print("Analysis complete!")
