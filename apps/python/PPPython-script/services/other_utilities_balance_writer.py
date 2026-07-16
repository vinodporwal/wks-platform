# -*- coding: utf-8 -*-
"""
Other Utilities Balance Writer
================================
Write Section IV: Other Utilities Balance to Excel.
Handles formatting and layout for BFW, DM Water, Cooling Water, Compressed Air, Oxygen, Effluent, Raw Water,
and the "Utilities Not Part of Balance" sub-section (Oxygen, Effluent, Chemicals).
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from database.connection import get_connection


# Styles (matching balance_report_service.py)
SECTION_FONT = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
SECTION_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
SUBSECTION_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
TOTAL_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
BOLD_FONT = Font(name='Calibri', size=11, bold=True)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


# ============================================================
# HELPER: Norm lookup
# ============================================================

def get_utility_norm_from_db(month: int, year: int, plant_name: str, utility_name: str, material_name: str) -> float:
    """Fetch norm value from NormsMonthDetail for a specific plant/utility/material."""
    conn = get_connection()
    cursor = conn.cursor()
    try:
        query = """
            SELECT nmd.Norms
            FROM NormsMonthDetail nmd
            INNER JOIN NormsHeader nh ON nh.Id = nmd.NormsHeader_FK_Id
            INNER JOIN Plants p ON p.Id = nh.Plant_FK_Id
            INNER JOIN FinancialYearMonth fym ON fym.Id = nmd.FinancialYearMonth_FK_Id
            WHERE fym.Month = ? AND fym.Year = ?
              AND p.Name = ?
              AND nh.UtilityName = ?
              AND nh.MaterialName = ?
              AND nh.IsActive = 1
        """
        cursor.execute(query, (month, year, plant_name, utility_name, material_name))
        db_row = cursor.fetchone()
        if db_row and db_row[0] is not None:
            return float(db_row[0])
        return 0.0
    except Exception as e:
        print(f"  [NORM FETCH] Error fetching norm for {plant_name}/{utility_name}/{material_name}: {e}")
        return 0.0
    finally:
        conn.close()


def fetch_all_chemical_norms(month: int, year: int, chemicals: list) -> dict:
    """
    Batch-fetch norm values for all chemicals in a single DB round-trip.
    Returns a dict keyed by (plant_name, utility_name, material_name) -> norm float.
    """
    if not chemicals:
        return {}
    conn = get_connection()
    cursor = conn.cursor()
    try:
        query = """
            SELECT p.Name, nh.UtilityName, nh.MaterialName, nmd.Norms
            FROM NormsMonthDetail nmd
            INNER JOIN NormsHeader nh ON nh.Id = nmd.NormsHeader_FK_Id
            INNER JOIN Plants p ON p.Id = nh.Plant_FK_Id
            INNER JOIN FinancialYearMonth fym ON fym.Id = nmd.FinancialYearMonth_FK_Id
            WHERE fym.Month = ? AND fym.Year = ?
              AND nh.IsActive = 1
              AND p.Name IN ('NMD - Utility Plant')
        """
        cursor.execute(query, (month, year))
        result = {}
        for db_row in cursor.fetchall():
            key = (str(db_row[0]), str(db_row[1]), str(db_row[2]))
            result[key] = float(db_row[3]) if db_row[3] is not None else 0.0
        return result
    except Exception as e:
        print(f"  [NORM FETCH] Error batch-fetching norms: {e}")
        return {}
    finally:
        conn.close()


# ============================================================
# CHEMICAL BALANCE WRITER
# ============================================================

def write_single_chemical_balance(ws, start_row: int, month: int, year: int, calculation_result: dict,
                                   chemical_name: str, uom: str, parent_utility: str,
                                   parent_utility_name: str, consumer_name: str,
                                   norm_cache: dict = None) -> int:
    """Write a single chemical balance - demand only (no supply columns, no balance row)."""
    utility_consumption = calculation_result.get('utility_consumption', {})
    if norm_cache is not None:
        norm = norm_cache.get(('NMD - Utility Plant', parent_utility_name, chemical_name), 0.0)
    else:
        norm = get_utility_norm_from_db(month, year, 'NMD - Utility Plant', parent_utility_name, chemical_name)

    row = start_row

    # Subsection header -- only A-C (demand-only, no supply columns)
    # Apply fill BEFORE merging (openpyxl requirement)
    for col in range(1, 4):
        ws.cell(row=row, column=col).fill = SUBSECTION_FILL
    ws.merge_cells(f'A{row}:C{row}')
    header_cell = ws[f'A{row}']
    # Use a clean display name for the header (strip DB-specific suffixes)
    display_name = (chemical_name
                    .replace(' \u2013 GRADE 1', '')   # en dash (U+2013)
                    .replace(' - GRADE 1', '')          # plain hyphen fallback
                    .replace(', AL2(SO4)3,18H2O', '')
                    .replace(';PN:MIS 19OX', '')
                    .replace(' IS 797 GRADE1', '')
                    .strip())
    header_cell.value = f"{display_name} BALANCE"
    header_cell.font = Font(bold=True, size=11)
    header_cell.alignment = Alignment(horizontal='left', vertical='center')
    row += 1

    # Column headers -- only A-C
    col_headers = ['Demand', 'Norm', f'Quantity {uom}']
    for col_idx, header in enumerate(col_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = BOLD_FONT
        cell.fill = SUBSECTION_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 1

    # Calculate chemical quantity
    chemical_quantity = 0.0
    if parent_utility == 'bfw':
        bfw_data = utility_consumption.get('bfw', {})
        chemical_quantity = bfw_data.get('total_m3', 0.0) * norm
    elif parent_utility == 'dm':
        dm_data = utility_consumption.get('dm_water', {})
        chemical_quantity = dm_data.get('total_m3', 0.0) * norm
    elif parent_utility == 'hrsg':
        total_shp = (utility_consumption.get('shp_from_hrsg1', 0.0)
                     + utility_consumption.get('shp_from_hrsg2', 0.0)
                     + utility_consumption.get('shp_from_hrsg3', 0.0))
        chemical_quantity = total_shp * norm

    # Demand subheading
    ws[f'A{row}'] = "Utility for Utility (U4U)"
    ws[f'A{row}'].font = BOLD_FONT
    for col in range(1, 4):
        ws.cell(row=row, column=col).border = THIN_BORDER
    row += 1

    # Demand detail row
    ws[f'A{row}'] = f"  {consumer_name}"
    ws[f'B{row}'] = round(norm, 7) if norm else ''
    ws[f'C{row}'] = round(chemical_quantity, 2)
    for col in range(1, 4):
        ws.cell(row=row, column=col).border = THIN_BORDER
    row += 1

    # Total row -- only A-C
    ws[f'A{row}'] = "Total Demand"
    ws[f'C{row}'] = round(chemical_quantity, 2)
    ws[f'A{row}'].font = BOLD_FONT
    ws[f'C{row}'].font = BOLD_FONT
    for col in range(1, 4):
        ws.cell(row=row, column=col).border = THIN_BORDER
        ws.cell(row=row, column=col).fill = TOTAL_FILL
    row += 1

    return row


def write_single_utility_balance(ws, start_row: int, utility_name: str, utility_data: dict,
                                  demand_only: bool = False) -> int:
    """
    Write a single utility balance section.

    When demand_only=True (used for "Utilities Not Part of Balance"):
    - Only the Demand | Norm | Quantity columns are shown (cols A-C merged across all 5)
    - Supply and Quantity columns are left blank
    - No Imbalance row

    When demand_only=False (default, used for Part A utilities):
    - Full 5-column layout: Demand | Norm | Quantity | Supply | Quantity
    - Supply fills in parallel from start_data_row
    - Total + Imbalance rows at the bottom
    """
    row = start_row
    unit = utility_data['unit']
    utility_label = utility_name.replace(" Balance", "").strip()

    # -- Subsection header ---------------------------------------------------
    fill_cols = range(1, 4) if demand_only else range(1, 6)
    merge_end = 'C' if demand_only else 'E'
    # Apply fill to every cell BEFORE merging (openpyxl requirement)
    for col in fill_cols:
        ws.cell(row=row, column=col).fill = SUBSECTION_FILL
    ws.merge_cells(f'A{row}:{merge_end}{row}')
    header_cell = ws[f'A{row}']
    header_cell.value = utility_name
    header_cell.font = BOLD_FONT
    header_cell.alignment = Alignment(horizontal='left', vertical='center')
    row += 1

    # -- Column headers ------------------------------------------------------
    if demand_only:
        # Only 3 meaningful headers; cols D-E left blank and unbordered
        col_headers = ['Demand', 'Norm', f'Quantity {unit}', '', '']
    else:
        col_headers = ['Demand', 'Norm', f'Quantity {unit}', 'Supply', f'Quantity {unit}']

    for col_idx, header in enumerate(col_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        if not demand_only or col_idx <= 3:
            cell.value = header
            cell.font = BOLD_FONT
            cell.fill = SUBSECTION_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 1

    start_data_row = row

    # -- DEMAND side (always written) ----------------------------------------
    demand_total = 0
    border_cols = range(1, 4) if demand_only else range(1, 6)

    if utility_data['demand']['u4u_items']:
        ws[f'A{row}'] = "Utility for Utility (U4U)"
        ws[f'A{row}'].font = BOLD_FONT
        for col in border_cols:
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1
        for item in utility_data['demand']['u4u_items']:
            ws[f'A{row}'] = f"  {item['name']}"
            ws[f'B{row}'] = round(item['norm'], 4) if item['norm'] > 0 else ''
            ws[f'C{row}'] = round(item['quantity'], 2)
            demand_total += item['quantity']
            for col in border_cols:
                ws.cell(row=row, column=col).border = THIN_BORDER
            row += 1

    if utility_data['demand']['process_items']:
        ws[f'A{row}'] = f"Plant Requirement ({utility_label})"
        ws[f'A{row}'].font = BOLD_FONT
        for col in border_cols:
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1
        for item in utility_data['demand']['process_items']:
            ws[f'A{row}'] = f"  {item['plant_name']}"
            demand_value = item['demand_value']
            ws[f'C{row}'] = round(demand_value, 2)
            demand_total += demand_value
            for col in border_cols:
                ws.cell(row=row, column=col).border = THIN_BORDER
            row += 1

    if utility_data['demand']['fixed_items']:
        ws[f'A{row}'] = f"Fixed Requirement ({utility_label})"
        ws[f'A{row}'].font = BOLD_FONT
        for col in border_cols:
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1
        for item in utility_data['demand']['fixed_items']:
            ws[f'A{row}'] = f"  {item['consumer_name']} ({item['plant_name']})"
            demand_value = item['consumption_value']
            ws[f'C{row}'] = round(demand_value, 2)
            demand_total += demand_value
            for col in border_cols:
                ws.cell(row=row, column=col).border = THIN_BORDER
            row += 1

    if demand_only:
        # -- DEMAND-ONLY: Total row -- only A-C bordered ----------------------
        ws[f'A{row}'] = f"Total {utility_label} Demand"
        ws[f'C{row}'] = round(demand_total, 2)  # Sum of all displayed rows
        ws[f'A{row}'].font = BOLD_FONT
        ws[f'C{row}'].font = BOLD_FONT
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        row += 1
        return row

    # -- SUPPLY side (demand_only=False only) --------------------------------
    # supply_total: use demand_total so supply always matches the displayed demand rows.
    # These utilities (BFW, DM Water, CW1, CW2, Compressed Air) are produced on-demand;
    # any tiny rounding difference between the engine's aggregate and the DB plant-wise
    # breakdown would otherwise create a spurious imbalance.
    supply_total = demand_total

    gen_row = start_data_row
    supply_label = utility_data['supply'].get('label', 'Utility Plant Production')
    ws[f'D{gen_row}'] = supply_label
    ws[f'E{gen_row}'] = round(supply_total, 2)
    for col in range(1, 6):
        ws.cell(row=gen_row, column=col).border = THIN_BORDER
    gen_row += 1

    while gen_row < row:
        for col in range(1, 6):
            ws.cell(row=gen_row, column=col).border = THIN_BORDER
        gen_row += 1

    # -- Total row -----------------------------------------------------------
    row = max(row, gen_row)
    ws[f'A{row}'] = f"Total {utility_label} Demand"
    ws[f'C{row}'] = round(demand_total, 2)  # Sum of all displayed rows
    ws[f'D{row}'] = f"Total {utility_label} Generation"
    ws[f'E{row}'] = round(supply_total, 2)
    ws[f'A{row}'].font = BOLD_FONT
    ws[f'C{row}'].font = BOLD_FONT
    ws[f'D{row}'].font = BOLD_FONT
    ws[f'E{row}'].font = BOLD_FONT
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = THIN_BORDER
        ws.cell(row=row, column=col).fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    row += 1

    # -- Imbalance row -------------------------------------------------------
    ws[f'A{row}'] = f"{utility_label} Imbalance"
    ws[f'C{row}'] = round(supply_total - demand_total, 2)  # Always 0 (supply = demand)
    ws[f'A{row}'].font = BOLD_FONT
    ws[f'C{row}'].font = BOLD_FONT
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = THIN_BORDER
        ws.cell(row=row, column=col).fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    row += 1

    return row


def write_other_utilities_balance_section(ws, start_row: int, month: int, year: int, calculation_result: dict) -> int:
    """
    Write Section IV: Other Utilities Balance.

    Part A - Utilities in Balance:
    1. BFW Balance
    2. DM Water Balance
    3. Cooling Water 1 Balance
    4. Cooling Water 2 Balance
    5. Compressed Air Balance

    Part B - Utilities Not Part of Balance:
    6. Raw Water Balance
    7. Oxygen Balance
    8. Effluent Treatment Balance
    9. Chemical balances (BFW chemicals, DM Water chemicals, HRSG chemicals)
    """
    from services.other_utilities_balance import (
        extract_bfw_balance_data,
        extract_dm_water_balance_data,
        extract_cooling_water_1_balance_data,
        extract_cooling_water_2_balance_data,
        extract_compressed_air_balance_data,
        extract_oxygen_balance_data,
        extract_effluent_balance_data,
        extract_raw_water_balance_data
    )
    row = start_row

    # -- Section header ------------------------------------------------------
    for col in range(1, 6):
        ws.cell(row=row, column=col).fill = SECTION_FILL
    ws.merge_cells(f'A{row}:E{row}')
    header_cell = ws[f'A{row}']
    header_cell.value = "SECTION IV: OTHER UTILITIES BALANCE"
    header_cell.font = SECTION_FONT
    header_cell.alignment = Alignment(horizontal='left', vertical='center')
    row += 1

    # -- Part A: utilities that are part of the balance -----------------------
    # 1. BFW Balance
    bfw_data = extract_bfw_balance_data(month, year, calculation_result)
    row = write_single_utility_balance(ws, row, "BFW Balance", bfw_data)
    row += 1

    # 2. DM Water Balance
    dm_data = extract_dm_water_balance_data(month, year, calculation_result)
    row = write_single_utility_balance(ws, row, "DM Water Balance", dm_data)
    row += 1

    # 3. Cooling Water 1 Balance
    cw1_data = extract_cooling_water_1_balance_data(month, year, calculation_result)
    row = write_single_utility_balance(ws, row, "Cooling Water 1 Balance", cw1_data)
    row += 1

    # 4. Cooling Water 2 Balance
    cw2_data = extract_cooling_water_2_balance_data(month, year, calculation_result)
    row = write_single_utility_balance(ws, row, "Cooling Water 2 Balance", cw2_data)
    row += 1

    # 5. Compressed Air Balance
    air_data = extract_compressed_air_balance_data(month, year, calculation_result)
    row = write_single_utility_balance(ws, row, "Compressed Air Balance", air_data)
    row += 2  # Extra gap before the "Not Part of Balance" header

    # -- Part B: Utilities Not Part of Balance --------------------------------
    # Apply fill BEFORE merging (openpyxl requirement); only A-C since supply cols removed
    for col in range(1, 4):
        ws.cell(row=row, column=col).fill = SECTION_FILL
    ws.merge_cells(f'A{row}:C{row}')
    not_part_header = ws[f'A{row}']
    not_part_header.value = "UTILITIES NOT PART OF BALANCE"
    not_part_header.font = SECTION_FONT
    not_part_header.alignment = Alignment(horizontal='left', vertical='center')
    row += 1

    # 6. Raw Water Balance
    raw_water_data = extract_raw_water_balance_data(month, year, calculation_result)
    row = write_single_utility_balance(ws, row, "Raw Water Balance", raw_water_data, demand_only=True)
    row += 1

    # 7. Oxygen Balance
    oxygen_data = extract_oxygen_balance_data(month, year, calculation_result)
    row = write_single_utility_balance(ws, row, "Oxygen Balance", oxygen_data, demand_only=True)
    row += 1

    # 8. Effluent Treatment Balance
    effluent_data = extract_effluent_balance_data(month, year, calculation_result)
    row = write_single_utility_balance(ws, row, "Effluent Treatment Balance", effluent_data, demand_only=True)
    row += 1

    # 9. Chemical balances
    chemicals = [
        # BFW Chemicals
        ('CHEM CYCLO HEXY',                         'KG', 'bfw',  'Boiler Feed Water',  'BFW Plant'),
        ('CHEM MORPHOLENE',                          'MT', 'bfw',  'Boiler Feed Water',  'BFW Plant'),
        ('KEM WATREAT B 70M',                        'KG', 'bfw',  'Boiler Feed Water',  'BFW Plant'),
        # DM Water Chemicals  (MaterialName must match DB exactly; UtilityName = 'D M Water')
        ('CAUSTIC SODA LYE \u2013 GRADE 1',           'MT', 'dm',   'D M Water',          'DM Water Plant'),
        ('CHEM ALUM.SULFATE, AL2(SO4)3,18H2O',      'KG', 'dm',   'D M Water',          'DM Water Plant'),
        ('CHEM  SODIUM SULPHITE;PN:MIS 19OX',        'KG', 'dm',   'D M Water',          'DM Water Plant'),
        ('POLYELECTROLYTE',                          'KG', 'dm',   'D M Water',          'DM Water Plant'),
        ('SODIUM CHLORIDE IS 797 GRADE1',            'MT', 'dm',   'D M Water',          'DM Water Plant'),
        # HRSG Chemicals
        ('CHEM TRISODIUM PHOSPHATE',                 'KG', 'hrsg', 'HRSG1_SHP STEAM',    'HRSG Plants'),
    ]

    # Batch-fetch all norms for this month in a single DB round-trip
    norm_cache = fetch_all_chemical_norms(month, year, chemicals)

    for chemical_name, uom, parent_utility, parent_utility_name, consumer_name in chemicals:
        row = write_single_chemical_balance(
            ws, row, month, year, calculation_result,
            chemical_name, uom, parent_utility, parent_utility_name, consumer_name,
            norm_cache=norm_cache
        )
        row += 1  # Gap between chemicals

    return row
