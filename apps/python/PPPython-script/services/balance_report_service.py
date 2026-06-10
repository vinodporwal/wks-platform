"""
Balance Report Service
======================
Generates Power & Utility Loop Balance Summary Excel Report.

This service creates a comprehensive monthly Excel report showing:
- Section I: Fuel Demand (GTs, STG, HRSGs with MMBTU calculations)
- Section II: Power Balance (detailed demand/supply breakdown)
- Section III: Steam Balance (SHP, HP, MP, LP with demand/supply)
- Section IV+: Other Utilities Balance (BFW, DM Water, CW, Air, Oxygen, Effluent)

The report is generated alongside calculation logs and saved in the same directory.
"""

import os
from datetime import datetime
from typing import Dict, List, Tuple, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from database.connection import get_connection
from database.power_asset_queries import get_month_hours, get_asset_capacity_limits, get_hrsg_heat_rate_for_load
from services.norm_lookup_service import get_month_norm


# ============================================================
# CONSTANTS
# ============================================================
MONTH_NAMES = {
    1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr",
    5: "May", 6: "Jun", 7: "Jul", 8: "Aug",
    9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"
}

# Excel styling - Consistent color scheme across all sections
SECTION_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Blue for section headers
SUBSECTION_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # Light blue for subsection headers
TOTAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Yellow for totals
FAILURE_FILL = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
FAILURE_DETAIL_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

SECTION_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")  # White text for section headers
BOLD_FONT = Font(name="Calibri", size=11, bold=True)
NORMAL_FONT = Font(name="Calibri", size=11)
FAILURE_FONT = Font(name="Calibri", size=12, bold=True, color="FFFFFF")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


# ============================================================
# HELPER FUNCTIONS FOR ASSET DEFAULT VALUES
# ============================================================

# Hardcoded fallback defaults (used only if DB fetch fails)
_FALLBACK_FIXED_MIN = {
    'GT1': 5.0, 'GT2': 5.0, 'GT3': 5.0, 'STG': 6.0
}
_FALLBACK_FIXED_MAX = {
    'GT1': 22.0, 'GT2': 22.0, 'GT3': 22.0, 'STG': 25.0
}


def get_asset_capacity_for_month(asset_name: str, month: int, year: int) -> dict:
    """
    Get month-wise MinOperatingCapacity and MaxOperatingCapacity for an asset from AssetAvailability table.
    Falls back to FixedMin and FixedMax if monthly values are not set.
    Same source as the Java /asset-capacity endpoint (CPP_NMD_GetAssetCapacity SP).
    
    Args:
        asset_name: Asset name (GT1, GT2, GT3, STG)
        month: Month number
        year: Year
        
    Returns:
        dict with 'MinCapacity' and 'MaxCapacity' (float or fallback)
    """
    try:
        limits = get_asset_capacity_limits(asset_name, month, year)
        
        # Try to get month-wise MinOperatingCapacity, fallback to FixedMin
        min_cap = limits.get('MinOperatingCapacity')
        if min_cap is None:
            min_cap = limits.get('FixedMin')
            
        # Try to get month-wise MaxOperatingCapacity, fallback to FixedMax
        max_cap = limits.get('MaxOperatingCapacity')
        if max_cap is None:
            max_cap = limits.get('FixedMax')
            
        if min_cap is not None and max_cap is not None:
            return {'MinCapacity': float(min_cap), 'MaxCapacity': float(max_cap)}
    except Exception as e:
        print(f"Warning: Could not get Min/Max Capacity for {asset_name}: {e}")
    
    # Fallback to hardcoded defaults
    return {
        'MinCapacity': _FALLBACK_FIXED_MIN.get(asset_name, 5.0),
        'MaxCapacity': _FALLBACK_FIXED_MAX.get(asset_name, 22.0)
    }


def get_asset_default_min_load(asset_name: str, month: int, year: int) -> float:
    """Get default min load for an asset."""
    return get_asset_capacity_for_month(asset_name, month, year)['MinCapacity']


def get_asset_default_max_load(asset_name: str, month: int, year: int) -> float:
    """Get default max load for an asset."""
    return get_asset_capacity_for_month(asset_name, month, year)['MaxCapacity']


# ============================================================
# DATA EXTRACTION FUNCTIONS
# ============================================================

def get_plant_wise_demand(month: int, year: int, utility_name: str) -> List[Dict]:
    """
    Fetch plant-wise demand from CalculatedProcessDemand table.
    Returns ALL NMD plants (using Power_Dis as master list) with 0 for any
    plant that has no demand for the requested utility/month.

    Args:
        month: Month number (1-12)
        year: Year
        utility_name: Utility name (e.g., 'Power_Dis', 'SHP Steam_Dis')

    Returns:
        List of dicts with plant_name and demand_value (0 if no data)
    """
    conn = get_connection()
    cursor = conn.cursor()

    try:
        # Get financial year string
        if month >= 4:
            fy_string = f"{year}-{str(year + 1)[-2:]}"
        else:
            fy_string = f"{year - 1}-{str(year)[-2:]}"

        # Month column mapping
        month_columns = {
            1: "jan", 2: "feb", 3: "mar", 4: "apr", 5: "may", 6: "jun",
            7: "jul", 8: "aug", 9: "sep", 10: "oct", 11: "nov", 12: "dec"
        }
        month_col = month_columns.get(month, "apr")

        # Use Power_Dis as the master plant list (always has all NMD plants).
        # LEFT JOIN to get this utility's value for the month; 0 if absent.
        query = f"""
            SELECT
                master.process_plant,
                ISNULL(target.{month_col}, 0) AS demand_value
            FROM (
                SELECT DISTINCT process_plant
                FROM dbo.CalculatedProcessDemand
                WHERE financial_year = ? AND cpp_utility = 'Power_Dis'
            ) AS master
            LEFT JOIN (
                SELECT process_plant, {month_col}
                FROM dbo.CalculatedProcessDemand
                WHERE financial_year = ? AND cpp_utility = ?
            ) AS target ON master.process_plant = target.process_plant
            ORDER BY master.process_plant
        """

        cursor.execute(query, (fy_string, fy_string, utility_name))
        rows = cursor.fetchall()

        result = []
        for row in rows:
            result.append({
                'plant_name': row[0],
                'demand_value': float(row[1])
            })

        return result

    except Exception as e:
        print(f"  [BALANCE REPORT] Error fetching plant-wise demand for {utility_name}: {e}")
        return []
    finally:
        conn.close()


def get_fixed_consumption_details(month: int, year: int, utility_name: str) -> List[Dict]:
    """
    Fetch fixed consumption with consumer and plant details.
    Returns ALL cost centers (using CPPCostCenters as master list) with 0 for any
    cost center that has no entry for the requested utility/month.

    Args:
        month: Month number (1-12)
        year: Year
        utility_name: Utility name (e.g., 'Power_Dis', 'LP Steam_Dis')

    Returns:
        List of dicts with consumer_name, plant_name, and consumption_value (0 if no data)
    """
    conn = get_connection()
    cursor = conn.cursor()

    try:
        # Get FinancialYearMonth ID
        cursor.execute("""
            SELECT Id FROM FinancialYearMonth WHERE Month = ? AND Year = ?
        """, (month, year))
        fym_row = cursor.fetchone()

        if not fym_row:
            return []

        fym_id = fym_row[0]

        # Use ALL CPPCostCenters as master list (LEFT JOIN).
        # Any cost center absent from UtilityFixedConsumption for this utility/month
        # will appear with ConsumptionValue = 0.
        query = """
            SELECT
                master.CostCenterName AS consumer_name,
                master.plant_name,
                ISNULL(ufc.ConsumptionValue, 0) AS consumption_value
            FROM (
                SELECT cc.CostCenterId, cc.CostCenterName, pm.DisplayName AS plant_name
                FROM CPPCostCenters cc
                JOIN FixedConsumptionPlantMapping pm ON cc.Plant_FK_Id = pm.Id
            ) AS master
            LEFT JOIN (
                SELECT ufc.CostCenter_FK_Id, ufc.ConsumptionValue
                FROM UtilityFixedConsumption ufc
                JOIN NormParameters np ON ufc.NormParameter_FK_Id = np.Id
                WHERE ufc.FinancialYearMonth_FK_Id = ? AND np.Name = ?
            ) AS ufc ON master.CostCenterId = ufc.CostCenter_FK_Id
            ORDER BY master.plant_name, master.CostCenterName
        """

        cursor.execute(query, (fym_id, utility_name))
        rows = cursor.fetchall()

        result = []
        for row in rows:
            result.append({
                'consumer_name': row[0],
                'plant_name': row[1] if row[1] else 'Unknown',
                'consumption_value': float(row[2]) if row[2] else 0.0
            })

        return result

    except Exception as e:
        print(f"  [BALANCE REPORT] Error fetching fixed consumption for {utility_name}: {e}")
        import traceback
        traceback.print_exc()
        return []
    finally:
        conn.close()


def get_ncv_from_fuel_availability(month: int, year: int, fuel_name: str = 'NATURAL GAS') -> float:
    """
    Fetch NCV (Net Calorific Value) from CPPFuelAvailability table for a specific month.
    
    Args:
        month: Month number (1-12)
        year: Year
        fuel_name: Fuel name (default: 'NATURAL GAS')
    
    Returns:
        NCV value in GBT (Gross Billion Therms) or 0 if not found
    """
    conn = get_connection()
    cursor = conn.cursor()
    
    try:
        # Get financial year string
        if month >= 4:
            fy_string = f"{year}-{str(year + 1)[-2:]}"
        else:
            fy_string = f"{year - 1}-{str(year)[-2:]}"
        
        # Month column mapping
        month_columns = {
            1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
            7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"
        }
        month_col = month_columns.get(month, "Apr")
        
        query = f"""
            SELECT {month_col}
            FROM dbo.CPPFuelAvailability
            WHERE FinancialYear = ? AND FuelName = ?
        """
        
        cursor.execute(query, (fy_string, fuel_name))
        row = cursor.fetchone()
        
        if row and row[0]:
            return float(row[0])
        return 0.0
        
    except Exception as e:
        print(f"Error fetching NCV from CPPFuelAvailability: {e}")
        return 0.0
    finally:
        cursor.close()
        conn.close()


def get_gt_heat_rate_and_free_steam(month: int, year: int, asset_name: str, gt_load_mw: float) -> tuple:
    """
    Fetch GT Heat Rate and Free Steam Factor from CPP_GTHeatRate table based on allocated load.
    Uses linear interpolation between nearest GTLoad points (same logic as power dispatch).
    
    Args:
        month: Month number (1-12)
        year: Year
        asset_name: Asset name (e.g., 'GT-1', 'GT-2', 'GT-3')
        gt_load_mw: GT allocated load in MW
    
    Returns:
        Tuple of (heat_rate, free_steam_factor) or (0.0, 0.0) if not found
    """
    conn = get_connection()
    cursor = conn.cursor()
    
    try:
        # Get financial year string
        if month >= 4:
            fy_string = f"{year}-{str(year + 1)[-2:]}"
        else:
            fy_string = f"{year - 1}-{str(year)[-2:]}"
        
        # Query heat rate table - get all load points for this specific GT asset, ordered by load
        query = """
            SELECT GTLoad, FinalHeatRate, FreeSteamFactor
            FROM dbo.CPP_GTHeatRate
            WHERE FinancialYear = ? AND AssetName = ?
            ORDER BY GTLoad
        """
        
        cursor.execute(query, (fy_string, asset_name))
        rows = cursor.fetchall()
        
        if not rows:
            return 0.0, 0.0

        # Normalize and sort points
        points = sorted(
            [(float(r[0]), float(r[1]), float(r[2])) for r in rows],
            key=lambda x: x[0]
        )

        if gt_load_mw is None:
            return 0.0, 0.0

        min_load = points[0][0]
        max_load = points[-1][0]

        # Clamp outside range to nearest endpoint (same as power_service behavior)
        if gt_load_mw <= min_load:
            _, hr, fs = points[0]
            return hr, fs
        if gt_load_mw >= max_load:
            _, hr, fs = points[-1]
            return hr, fs

        # Exact match or find interpolation bracket
        lower = None
        upper = None
        for load_point, heat_rate, free_steam in points:
            if abs(load_point - gt_load_mw) < 1e-9:
                return heat_rate, free_steam
            if load_point < gt_load_mw:
                lower = (load_point, heat_rate, free_steam)
            elif load_point > gt_load_mw and upper is None:
                upper = (load_point, heat_rate, free_steam)
                break

        if lower is None or upper is None:
            return 0.0, 0.0

        x1, hr1, fs1 = lower
        x2, hr2, fs2 = upper

        if abs(x2 - x1) < 1e-9:
            return hr1, fs1

        frac = (gt_load_mw - x1) / (x2 - x1)
        heat = hr1 + frac * (hr2 - hr1)
        steam = fs1 + frac * (fs2 - fs1)

        # If interpolated heat rate is 0, use nearest non-zero point
        if heat == 0.0:
            min_diff = float('inf')
            nearest_heat = 0.0
            nearest_steam = 0.0
            for load_point, row_heat, row_steam in points:
                if row_heat != 0.0:
                    diff = abs(load_point - gt_load_mw)
                    if diff < min_diff:
                        min_diff = diff
                        nearest_heat = row_heat
                        nearest_steam = row_steam
            heat = nearest_heat
            steam = nearest_steam

        return heat, steam
        
    except Exception as e:
        print(f"Error fetching GT Heat Rate and Free Steam: {e}")
        return 0.0, 0.0
    finally:
        cursor.close()
        conn.close()


def extract_asset_availability_data(month: int, year: int, calculation_result: dict) -> Dict:
    """
    Extract asset availability and loading data for all generation assets.
    
    Returns dict with:
    - generation_assets: List of power generation assets (GTs, STG, Import)
    - steam_assets: List of steam generation assets (HRSGs)
    """
    usd_result = calculation_result.get('usd_result', {})
    final_dispatch = usd_result.get('final_dispatch', [])
    power_result = usd_result.get('power_result', {})
    
    # Get HRSG availability from steam calculation
    from services.steam_service import get_hrsg_availability_from_dispatch
    hrsg_availability = {}
    if final_dispatch:
        hrsg_availability = get_hrsg_availability_from_dispatch(final_dispatch)
    
    # Extract generation assets (GTs, STG, Import Power)
    generation_assets = []
    
    # Process GTs
    gt_names = ['GT1', 'GT2', 'GT3']
    
    # Debug: Print all available asset names in dispatch
    print(f"[DEBUG] Available assets in dispatch: {[asset.get('AssetName', 'Unknown') for asset in final_dispatch]}")
    
    for gt_name in gt_names:
        gt_data = None
        for asset in final_dispatch:
            asset_name = asset.get('AssetName', '')
            # More flexible matching for GT assets - check multiple possible patterns
            if (gt_name in asset_name.upper() or 
                f'Plant-{gt_name[-1]}' in asset_name or 
                f'Plant {gt_name[-1]}' in asset_name or
                f'GT{gt_name[-1]}' in asset_name.upper() or
                (gt_name == 'GT1' and ('GT1' in asset_name.upper() or '1' in asset_name and 'GT' in asset_name.upper())) or
                (gt_name == 'GT2' and ('GT2' in asset_name.upper() or '2' in asset_name and 'GT' in asset_name.upper())) or
                (gt_name == 'GT3' and ('GT3' in asset_name.upper() or '3' in asset_name and 'GT' in asset_name.upper()))):
                gt_data = asset
                print(f"[DEBUG] Found {gt_name} in dispatch: {asset_name}")
                break
        
        if gt_data:
            gross_mwh = gt_data.get('GrossMWh', 0)
            hours = gt_data.get('Hours', get_month_hours(month, year))
            load_mw = gt_data.get('LoadMW', 0)
            # Get Month-wise Min/Max from AssetAvailability table
            _gt_limits = get_asset_capacity_for_month(gt_name, month, year)
            min_mw = _gt_limits['MinCapacity']
            max_mw = _gt_limits['MaxCapacity']
            
            print(f"[DEBUG] {gt_name} dispatch data: MinMW={min_mw}, MaxMW={max_mw}, LoadMW={load_mw}")
            
            # If no LoadMW, calculate from GrossMWh and Hours
            if load_mw == 0 and hours > 0:
                load_mw = gross_mwh / hours
            
            generation_assets.append({
                'asset_name': gt_name,
                'availability_hr': hours if gross_mwh > 0 else 0,
                'min_capacity': min_mw,
                'max_capacity': max_mw,
                'avg_load_per_hr': round(load_mw, 2)
            })
        else:
            # GT not in dispatch (not running) - try to get default min load from database
            print(f"[DEBUG] {gt_name} not found in dispatch, using database defaults")
            default_min_mw = get_asset_default_min_load(gt_name, month, year)
            default_max_mw = get_asset_default_max_load(gt_name, month, year)
            
            generation_assets.append({
                'asset_name': gt_name,
                'availability_hr': 0,
                'min_capacity': default_min_mw,
                'max_capacity': default_max_mw,
                'avg_load_per_hr': 0
            })
    
    # Process STG
    stg_data = None
    for asset in final_dispatch:
        asset_name = asset.get('AssetName', '')
        if 'STG' in asset_name:
            stg_data = asset
            break
    
    if stg_data:
        gross_mwh = stg_data.get('GrossMWh', 0)
        hours = stg_data.get('Hours', get_month_hours(month, year))
        load_mw = stg_data.get('LoadMW', 0)
        _stg_limits = get_asset_capacity_for_month('STG', month, year)
        min_mw = _stg_limits['MinCapacity']
        max_mw = _stg_limits['MaxCapacity']
        
        if load_mw == 0 and hours > 0:
            load_mw = gross_mwh / hours
        
        generation_assets.append({
            'asset_name': 'STG',
            'availability_hr': hours if gross_mwh > 0 else 0,
            'min_capacity': min_mw,
            'max_capacity': max_mw,
            'avg_load_per_hr': round(load_mw, 2)
        })
    else:
        # STG not in dispatch - get default min/max load from database
        default_min_mw = 6.0  # Hardcoded STG min capacity as requested
        default_max_mw = get_asset_default_max_load('STG', month, year)
        
        generation_assets.append({
            'asset_name': 'STG',
            'availability_hr': 0,
            'min_capacity': default_min_mw,
            'max_capacity': default_max_mw,
            'avg_load_per_hr': 0
        })
    
    # Process Import Power
    import_mwh = power_result.get('importUnits', 0)
    import_hours = get_month_hours(month, year)  # Use actual month hours
    
    # Get Import asset capacity from final_dispatch
    import_data = None
    for asset in final_dispatch:
        asset_name = asset.get('AssetName', '')
        if 'IMPORT' in asset_name.upper() or 'GRID' in asset_name.upper():
            import_data = asset
            break
    
    import_max_mw = import_data.get('CapacityMW', 25) if import_data else 25
    import_load_mw = round(import_mwh / import_hours, 2) if import_hours > 0 and import_mwh > 0 else 0
    
    generation_assets.append({
        'asset_name': 'Import',
        'availability_hr': import_hours if import_mwh > 0 else 0,
        'min_capacity': 0,
        'max_capacity': import_max_mw,
        'avg_load_per_hr': import_load_mw
    })
    
    # Extract steam assets (HRSGs)
    hrsg_dispatch = usd_result.get('hrsg_dispatch', {})
    hrsg_dispatch_list = hrsg_dispatch.get('hrsg_dispatch', [])
    
    # Create map of total SHP (fired + free) from iteration results
    hrsg_total_shp_map = {}
    for h in hrsg_dispatch_list:
        hrsg_name = h.get('name', '')
        # Remove any spaces or dashes to match "HRSG1" format
        clean_name = hrsg_name.replace(' ', '').replace('-', '').upper()
        if 'HRSG1' in clean_name:
            hrsg_total_shp_map['HRSG1'] = h.get('total_shp_mt', 0)
        elif 'HRSG2' in clean_name:
            hrsg_total_shp_map['HRSG2'] = h.get('total_shp_mt', 0)
        elif 'HRSG3' in clean_name:
            hrsg_total_shp_map['HRSG3'] = h.get('total_shp_mt', 0)
    
    steam_assets = []
    hrsg_names = ['HRSG1', 'HRSG2', 'HRSG3']
    
    # Create GT to HRSG mapping
    gt_hrsg_map = {
        'HRSG1': 'GT1',
        'HRSG2': 'GT2',
        'HRSG3': 'GT3'
    }
    
    for hrsg_name in hrsg_names:
        # Get linked GT availability hours
        linked_gt = gt_hrsg_map.get(hrsg_name)
        gt_hours = 0
        
        # Find the linked GT's availability hours from generation_assets
        for gen_asset in generation_assets:
            if gen_asset['asset_name'] == linked_gt:
                gt_hours = gen_asset['availability_hr']
                break
        
        # Calculate total allocated load (fired + free)
        total_shp_mt = hrsg_total_shp_map.get(hrsg_name, 0)
        avg_total_load = total_shp_mt / gt_hours if gt_hours > 0 else 0
        
        display_name = f"{hrsg_name} (fired + free steam)"
        
        if hrsg_name in hrsg_availability:
            hrsg_data = hrsg_availability[hrsg_name]
            min_cap = hrsg_data.get('min_capacity_mt', 60)
            max_cap = hrsg_data.get('max_capacity_mt', 130)
            
            steam_assets.append({
                'asset_name': display_name,
                'availability_hr': gt_hours,
                'min_capacity': min_cap,
                'max_capacity': max_cap,
                'avg_load_per_hr': round(avg_total_load, 2)
            })
        else:
            steam_assets.append({
                'asset_name': display_name,
                'availability_hr': gt_hours,
                'min_capacity': 60,
                'max_capacity': 130,
                'avg_load_per_hr': round(avg_total_load, 2)
            })
    
    return {
        'generation_assets': generation_assets,
        'steam_assets': steam_assets
    }


def extract_fuel_demand_data(month: int, year: int, calculation_result: dict) -> Dict:
    """
    Extract fuel demand data for GTs, STG, and HRSGs.
    
    Args:
        month: Month number (1-12)
        year: Year
        calculation_result: Calculation result dictionary
    
    Returns dict with:
    - gt_assets: List of GT fuel consumption data
    - stg_asset: STG fuel consumption data
    - hrsg_assets: List of HRSG fuel consumption data
    """
    usd_result = calculation_result.get('usd_result', {})
    final_dispatch = usd_result.get('final_dispatch', [])
    utility_consumption = calculation_result.get('utility_consumption', {})
    natural_gas = utility_consumption.get('natural_gas', {})
    stg_extraction = calculation_result.get('stg_extraction', {})
    
    # Fetch NCV from database for this month
    ncv_gbt = get_ncv_from_fuel_availability(month, year, 'NATURAL GAS')
    
    # Get actual operating hours for the month based on days in month
    # May 2025 = 31 days = 744 hours, April 2025 = 30 days = 720 hours, etc.
    default_operating_hours = get_month_hours(month, year)
    
    # Get total SHP (fired + free) from hrsg_dispatch for accurate heat rate calculation
    hrsg_dispatch = usd_result.get('hrsg_dispatch', {})
    hrsg_dispatch_list = hrsg_dispatch.get('hrsg_dispatch', [])
    hrsg_total_shp_map = {}
    for h in hrsg_dispatch_list:
        hrsg_name = h.get('name', '')
        clean_name = hrsg_name.replace(' ', '').replace('-', '').upper()
        if 'HRSG1' in clean_name:
            hrsg_total_shp_map['HRSG1'] = h.get('total_shp_mt', 0)
        elif 'HRSG2' in clean_name:
            hrsg_total_shp_map['HRSG2'] = h.get('total_shp_mt', 0)
        elif 'HRSG3' in clean_name:
            hrsg_total_shp_map['HRSG3'] = h.get('total_shp_mt', 0)
    
    # Extract GT data - Create map from dispatch data
    gt_dispatch_map = {}
    for asset in final_dispatch:
        asset_name = asset.get('AssetName', '')
        if 'GT' in asset_name or 'Power Plant' in asset_name:
            gross_mwh = asset.get('GrossMWh', 0)
            operating_hours = asset.get('OperatingHours', default_operating_hours)
            
            if 'Plant-1' in asset_name or 'Plant 1' in asset_name:
                gt_dispatch_map['GT1'] = {'gross_mwh': gross_mwh, 'hours': operating_hours}
            elif 'Plant-2' in asset_name or 'Plant 2' in asset_name:
                gt_dispatch_map['GT2'] = {'gross_mwh': gross_mwh, 'hours': operating_hours}
            elif 'Plant-3' in asset_name or 'Plant 3' in asset_name:
                gt_dispatch_map['GT3'] = {'gross_mwh': gross_mwh, 'hours': operating_hours}
    
    # Build GT assets list - ALWAYS show all 3 GTs
    gt_configs = [
        ('GT1', 'GT-1', 'gt1_ng_details'),
        ('GT2', 'GT-2', 'gt2_ng_details'),
        ('GT3', 'GT-3', 'gt3_ng_details')
    ]
    
    gt_assets = []
    for gt_name, gt_db_name, ng_details_key in gt_configs:
        dispatch_data = gt_dispatch_map.get(gt_name, {'gross_mwh': 0, 'hours': default_operating_hours})
        gross_mwh = dispatch_data['gross_mwh']
        gross_kwh = gross_mwh * 1000
        operating_hours = dispatch_data['hours']
        avg_load_mw = gross_mwh / operating_hours if operating_hours > 0 else 0
        
        # If GT is off (gross_mwh = 0), show NCV from DB but zeroes for load/MMBTU/heat rate
        if gross_mwh == 0:
            gt_assets.append({
                'asset_name': gt_name,
                'ncv_kcal_kwh': ncv_gbt,  # Always show actual NCV even when GT is off
                'quantity_mmbtu': 0,
                'allocated_load_mw': 0,
                'gross_mwh': 0,
                'heat_rate_kcal_kwh': 0
            })
        else:
            # Calculate Net MMBTU from heat rate and free steam factor
            # Get heat rate and free steam factor from database based on actual load
            heat_rate_kcal_kwh, free_steam_factor = get_gt_heat_rate_and_free_steam(month, year, gt_db_name, avg_load_mw)
            
            print(f"[FUEL DEMAND DEBUG] {gt_name}: Load={avg_load_mw:.2f} MW, Heat Rate={heat_rate_kcal_kwh:.2f}, FreeSteam={free_steam_factor:.4f}")
            
            if heat_rate_kcal_kwh > 0:
                # Constants for calculation
                GT_NG_KCAL_TO_BTU = 3.96567
                GT_NG_BTU_TO_MMBTU = 1000000
                GT_NG_FREE_STEAM_ENERGY_KCAL_KG = 760.87
                
                # Calculate Gross MMBTU = KWH × Heat Rate × 3.96567 / 1,000,000
                gross_mmbtu = gross_kwh * heat_rate_kcal_kwh * GT_NG_KCAL_TO_BTU / GT_NG_BTU_TO_MMBTU
                
                # Calculate Free Steam MMBTU = KWH × FreeSteamFactor × 760.87 × 3.96567 / 1,000,000
                free_steam_mmbtu = gross_kwh * free_steam_factor * GT_NG_FREE_STEAM_ENERGY_KCAL_KG * GT_NG_KCAL_TO_BTU / GT_NG_BTU_TO_MMBTU
                
                # Calculate Net MMBTU = Gross - Free Steam
                quantity_mmbtu = gross_mmbtu - free_steam_mmbtu
                
                print(f"[FUEL DEMAND DEBUG] {gt_name}: Gross={gross_mmbtu:.2f}, FreeSteam={free_steam_mmbtu:.2f}, Net={quantity_mmbtu:.2f} MMBTU")
            else:
                # Fallback: Try to get from ng_details if available
                ng_details = natural_gas.get(ng_details_key)
                if ng_details and isinstance(ng_details, dict) and 'gross_mmbtu' in ng_details:
                    quantity_mmbtu = ng_details.get('gross_mmbtu', 0)
                    heat_rate_kcal_kwh = ng_details.get('heat_rate', 0)
                    print(f"[FUEL DEMAND DEBUG] {gt_name}: Using ng_details gross_mmbtu={quantity_mmbtu:.2f}")
                else:
                    # Last resort: use absolute value of net MMBTU
                    gt_mmbtu_key = f'{gt_name.lower()}_mmbtu'
                    net_mmbtu = natural_gas.get(gt_mmbtu_key, 0)
                    quantity_mmbtu = abs(net_mmbtu)
                    print(f"[FUEL DEMAND DEBUG] {gt_name}: FALLBACK - Using abs(net_mmbtu)={quantity_mmbtu:.2f}")
                    
                    # Calculate heat rate from norm
                    ng_norm = natural_gas.get(f'{gt_name.lower()}_ng_norm', 0.0095)
                    heat_rate_kcal_kwh = abs(ng_norm) * 251995.76
            
            gt_assets.append({
                'asset_name': gt_name,
                'ncv_kcal_kwh': ncv_gbt,  # Use NCV from database
                'quantity_mmbtu': quantity_mmbtu,
                'allocated_load_mw': avg_load_mw,
                'gross_mwh': gross_mwh,  # Total monthly generation
                'heat_rate_kcal_kwh': heat_rate_kcal_kwh  # Use heat rate from database
            })
    
    # Extract STG data - ALWAYS show STG even if not running
    stg_asset = {
        'asset_name': 'STG',
        'shp_consumed_mt': 0,
        'allocated_load_mw': 0,
        'gross_mwh': 0,
        'ncv_gbt': ncv_gbt  # Use same NCV as GTs
    }
    
    for asset in final_dispatch:
        asset_name = asset.get('AssetName', '')
        if 'STG' in asset_name:
            gross_mwh = asset.get('GrossMWh', 0)
            operating_hours = asset.get('OperatingHours', default_operating_hours)
            avg_load_mw = gross_mwh / operating_hours if operating_hours > 0 else 0
            
            # STG consumes SHP steam
            shp_inlet_mt = stg_extraction.get('stg_shp_inlet_mt', 0)
            
            stg_asset = {
                'asset_name': 'STG',
                'shp_consumed_mt': shp_inlet_mt,
                'allocated_load_mw': avg_load_mw,
                'gross_mwh': gross_mwh,
                'ncv_gbt': ncv_gbt  # Use same NCV as GTs
            }
            break
    
    # Extract HRSG data - ALWAYS show all 3 HRSGs
    # Get HRSG SHP generation from utility_consumption
    hrsg_assets = []
    hrsg_configs = [
        ('HRSG1', 'shp_from_hrsg1', 'hrsg1_mmbtu', natural_gas.get('hrsg1_ng_norm', 2.8115696), natural_gas.get('hrsg1_heat_rate', 0.0)),
        ('HRSG2', 'shp_from_hrsg2', 'hrsg2_mmbtu', natural_gas.get('hrsg2_ng_norm', 2.8115696), natural_gas.get('hrsg2_heat_rate', 0.0)),
        ('HRSG3', 'shp_from_hrsg3', 'hrsg3_mmbtu', natural_gas.get('hrsg3_ng_norm', 2.8115696), natural_gas.get('hrsg3_heat_rate', 0.0))
    ]
    
    gt_to_hrsg_map = {
        'HRSG1': 'GT1',
        'HRSG2': 'GT2',
        'HRSG3': 'GT3'
    }
    
    for hrsg_name, shp_key, mmbtu_key, ng_norm_per_mt, lookup_heat_rate_btu_lb in hrsg_configs:
        linked_gt = gt_to_hrsg_map.get(hrsg_name, '')
        gt_data = gt_dispatch_map.get(linked_gt, {'gross_mwh': 0, 'hours': default_operating_hours})
        total_shp_mt = hrsg_total_shp_map.get(hrsg_name, 0)

        # HRSG off when linked GT is not running or no total SHP output
        if gt_data['gross_mwh'] == 0 or total_shp_mt <= 0:
            hrsg_assets.append({
                'asset_name': hrsg_name,
                'ncv_gbt': ncv_gbt,
                'quantity_mmbtu': 0,
                'allocated_load_mt_per_hr': 0,
                'heat_rate_btu_lb': 0,
            })
            continue

        # Average hourly load (MT/h) from total SHP (fired + free steam)
        avg_load_mt_per_hr = total_shp_mt / default_operating_hours if default_operating_hours > 0 else 0

        quantity_mmbtu = natural_gas.get(mmbtu_key, 0)
        if quantity_mmbtu == 0 and total_shp_mt > 0:
            quantity_mmbtu = total_shp_mt * ng_norm_per_mt

        # Heat rate from CPP_HRSGHeatRate lookup at this load (BTU/lb) — same as Inputs screen
        heat_rate_btu_lb = 0.0
        try:
            hr_result = get_hrsg_heat_rate_for_load(
                equipment_name=hrsg_name,
                hrsg_load_tph=avg_load_mt_per_hr,
                month=month,
                year=year,
            )
            if hr_result.get('heat_rate_btu_lb', 0) > 0:
                heat_rate_btu_lb = hr_result['heat_rate_btu_lb']
            elif lookup_heat_rate_btu_lb and lookup_heat_rate_btu_lb > 0:
                heat_rate_btu_lb = lookup_heat_rate_btu_lb
        except Exception:
            if lookup_heat_rate_btu_lb and lookup_heat_rate_btu_lb > 0:
                heat_rate_btu_lb = lookup_heat_rate_btu_lb

        hrsg_assets.append({
            'asset_name': hrsg_name,
            'ncv_gbt': ncv_gbt,
            'quantity_mmbtu': quantity_mmbtu,
            'allocated_load_mt_per_hr': avg_load_mt_per_hr,
            'heat_rate_btu_lb': heat_rate_btu_lb,
        })
    
    return {
        'gt_assets': gt_assets,
        'stg_asset': stg_asset,
        'hrsg_assets': hrsg_assets
    }


def extract_power_balance_data(month: int, year: int, calculation_result: dict) -> Dict:
    """
    Extract complete power balance data with detailed demand and supply breakdown.
    """
    usd_result = calculation_result.get('usd_result', {})
    power_result = usd_result.get('power_result', {})
    final_dispatch = usd_result.get('final_dispatch', [])
    utility_consumption = calculation_result.get('utility_consumption', {})
    
    # DEMAND SIDE
    demand_data = {
        'plant_requirement': get_plant_wise_demand(month, year, 'Power_Dis'),
        'fixed_requirement': get_fixed_consumption_details(month, year, 'Power_Dis'),
        'u4u_requirement': [],
        'total_demand_mwh': power_result.get('totalDemandUnits', 0)
    }
    
    # U4U (Utility for Utility) - auxiliary consumption
    u4u_power = utility_consumption.get('u4u_power', {})
    for key, value in u4u_power.items():
        if value > 0:
            # Convert key to readable name
            readable_name = key.replace('_', ' ').title()
            demand_data['u4u_requirement'].append({
                'utility_name': readable_name,
                'consumption_mwh': value / 1000  # Convert KWH to MWH
            })
    
    # SUPPLY SIDE
    supply_data = {
        'assets': [],
        'total_supply_mwh': power_result.get('totalNetGeneration', 0) + power_result.get('importUnits', 0)
    }
    
    # Extract generation by asset
    for asset in final_dispatch:
        asset_name = asset.get('AssetName', '')
        net_mwh = asset.get('NetMWh', 0)
        
        if net_mwh > 0:
            supply_data['assets'].append({
                'asset_name': asset_name,
                'net_mwh': net_mwh
            })
    
    # Add import power
    import_mwh = power_result.get('mandatoryImportUsed', 0)
    if import_mwh > 0:
        supply_data['assets'].append({
            'asset_name': 'Import Power (MEL)',
            'net_mwh': import_mwh
        })
    
    # BALANCE
    balance = demand_data['total_demand_mwh'] - supply_data['total_supply_mwh']
    
    return {
        'demand': demand_data,
        'supply': supply_data,
        'balance_mwh': balance
    }


def extract_steam_balance_data(month: int, year: int, calculation_result: dict, steam_type: str) -> Dict:
    """
    Extract steam balance data for a specific steam type (SHP, HP, MP, LP).
    
    Args:
        steam_type: One of 'SHP', 'HP', 'MP', 'LP'
    """
    usd_result = calculation_result.get('usd_result', {})
    final_steam = usd_result.get('final_steam_balance', {})
    utility_consumption = calculation_result.get('utility_consumption', {})
    stg_extraction = calculation_result.get('stg_extraction', {})
    
    steam_type_lower = steam_type.lower()
    steam_balance = final_steam.get(f'{steam_type_lower}_balance', {})
    
    # Map steam type to utility name in database
    utility_name_map = {
        'SHP': 'SHP Steam_Dis',
        'HP': 'HP Steam_Dis',
        'MP': 'MP Steam_Dis',
        'LP': 'LP Steam_Dis'
    }
    utility_name = utility_name_map.get(steam_type, f'{steam_type} Steam_Dis')
    
    # DEMAND SIDE
    demand_data = {
        'plant_requirement': get_plant_wise_demand(month, year, utility_name),
        'fixed_requirement': get_fixed_consumption_details(month, year, utility_name),
        'stg_consumption': 0,
        'prds_consumption': 0,
        'u4u_requirement': [],
        'total_demand_mt': steam_balance.get(f'{steam_type_lower}_total', 0)
    }
    
    # STG consumption (for SHP only)
    if steam_type == 'SHP':
        demand_data['stg_consumption'] = stg_extraction.get('stg_shp_inlet_mt', 0)
    
    # PRDS consumption
    if steam_type == 'SHP':
        # SHP consumed by PRDS to generate HP, MP, LP
        demand_data['prds_consumption'] = steam_balance.get('shp_to_prds', 0)
    elif steam_type == 'MP':
        # MP consumed by PRDS to generate LP
        demand_data['prds_consumption'] = steam_balance.get('mp_to_prds', 0)
    
    # SUPPLY SIDE
    supply_data = {
        'sources': [],
        'total_supply_mt': steam_balance.get(f'{steam_type_lower}_total_supply', 0) or steam_balance.get(f'total_{steam_type_lower}_supply', 0)
    }
    
    # Extract supply sources based on steam type
    if steam_type == 'SHP':
        # SHP from HRSGs (free + supplementary)
        for i in [1, 2, 3]:
            shp_from_hrsg = utility_consumption.get(f'shp_from_hrsg{i}', 0)
            if shp_from_hrsg > 0:
                supply_data['sources'].append({
                    'source_name': f'HRSG{i}',
                    'supply_mt': shp_from_hrsg
                })
    
    elif steam_type == 'HP':
        # HP from PRDS (SHP → HP)
        hp_from_prds = steam_balance.get('hp_from_prds', 0)
        if hp_from_prds > 0:
            supply_data['sources'].append({
                'source_name': 'HP PRDS (from SHP)',
                'supply_mt': hp_from_prds
            })
    
    elif steam_type == 'MP':
        # MP from PRDS (SHP → MP)
        mp_from_prds = steam_balance.get('mp_from_prds', 0)
        if mp_from_prds > 0:
            supply_data['sources'].append({
                'source_name': 'MP PRDS (from SHP)',
                'supply_mt': mp_from_prds
            })
        
        # MP from STG extraction
        mp_from_stg = stg_extraction.get('mp_from_stg', 0)
        if mp_from_stg > 0:
            supply_data['sources'].append({
                'source_name': 'STG Extraction',
                'supply_mt': mp_from_stg
            })
    
    elif steam_type == 'LP':
        # LP from PRDS (MP → LP)
        lp_from_prds = steam_balance.get('lp_from_prds', 0)
        if lp_from_prds > 0:
            supply_data['sources'].append({
                'source_name': 'LP PRDS (from MP)',
                'supply_mt': lp_from_prds
            })
        
        # LP from STG extraction
        lp_from_stg = stg_extraction.get('lp_from_stg', 0)
        if lp_from_stg > 0:
            supply_data['sources'].append({
                'source_name': 'STG Extraction',
                'supply_mt': lp_from_stg
            })
        
        # LP credit from HRSG (negative - returned to BFW)
        lp_credit = steam_balance.get('lp_credit_from_hrsg', 0)
        if lp_credit != 0:
            supply_data['sources'].append({
                'source_name': 'HRSG LP Credit (Return)',
                'supply_mt': lp_credit
            })
    
    # BALANCE
    balance = demand_data['total_demand_mt'] - supply_data['total_supply_mt']
    
    return {
        'demand': demand_data,
        'supply': supply_data,
        'balance_mt': balance
    }


def extract_utility_balance_data(month: int, year: int, calculation_result: dict, utility_name: str, utility_key: str) -> Dict:
    """
    Extract balance data for other utilities (BFW, DM Water, CW, Air, Oxygen, Effluent).
    
    Args:
        utility_name: Display name (e.g., 'Boiler Feed Water')
        utility_key: Key in utility_consumption dict (e.g., 'bfw')
    """
    utility_consumption = calculation_result.get('utility_consumption', {})
    
    # Special handling for oxygen and effluent which are stored in utility_consumption
    if utility_key == 'oxygen':
        oxygen_mt = abs(utility_consumption.get('oxygen_mt', 0))  # Use absolute value
        demand_data = {
            'plant_requirement': get_plant_wise_demand(month, year, utility_name),
            'fixed_requirement': get_fixed_consumption_details(month, year, utility_name),
            'u4u_requirement': [],
            'total_demand': oxygen_mt
        }
        supply_data = {
            'total_supply': oxygen_mt
        }
        balance = 0.0
        unit = 'MT'
    elif utility_key == 'effluent':
        effluent_m3 = abs(utility_consumption.get('effluent_m3', 0))  # Use absolute value
        demand_data = {
            'plant_requirement': get_plant_wise_demand(month, year, utility_name),
            'fixed_requirement': get_fixed_consumption_details(month, year, utility_name),
            'u4u_requirement': [],
            'total_demand': effluent_m3
        }
        supply_data = {
            'total_supply': effluent_m3
        }
        balance = 0.0
        unit = 'M3'
    else:
        # Standard handling for other utilities
        utility_data = utility_consumption.get(utility_key, {})
        
        # DEMAND SIDE - use absolute values
        total_demand = utility_data.get('total_demand', 0) or utility_data.get('total', 0)
        demand_data = {
            'plant_requirement': get_plant_wise_demand(month, year, utility_name),
            'fixed_requirement': get_fixed_consumption_details(month, year, utility_name),
            'u4u_requirement': [],
            'total_demand': abs(total_demand)  # Use absolute value
        }
        
        # SUPPLY SIDE - use absolute values
        total_supply = utility_data.get('total_m3', 0) or utility_data.get('total_mt', 0) or utility_data.get('total_nm3', 0) or utility_data.get('total_km3', 0)
        supply_data = {
            'total_supply': abs(total_supply)  # Use absolute value
        }
        
        # BALANCE
        balance = demand_data['total_demand'] - supply_data['total_supply']
        unit = utility_data.get('unit', '')
    
    return {
        'demand': demand_data,
        'supply': supply_data,
        'balance': balance,
        'unit': unit
    }


def get_failure_details(calculation_result: dict) -> Optional[Dict]:
    balance_failure = calculation_result.get('balance_failure')
    if balance_failure:
        return {
            'message': balance_failure.get('message', 'Balance failure detected'),
            'suggestion': balance_failure.get('suggestion', 'Review asset capacities and balance drivers.')
        }

    for error in calculation_result.get('errors', []):
        if error.get('stage') == 'BALANCE_CHECK':
            return {
                'message': error.get('message', 'Balance failure detected'),
                'suggestion': error.get('suggestion', 'Review asset capacities and balance drivers.')
            }

    return None


def write_failure_banner(ws, start_row: int, calculation_result: dict) -> int:
    failure_details = get_failure_details(calculation_result)
    if not failure_details:
        return start_row

    for col in range(1, 6):
        ws.cell(row=start_row, column=col).fill = FAILURE_FILL
    ws.merge_cells(f'A{start_row}:E{start_row}')
    title_cell = ws[f'A{start_row}']
    title_cell.value = 'MONTH FAILED - BALANCE CHECK FAILED'
    title_cell.font = FAILURE_FONT
    title_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    start_row += 1

    for col in range(1, 6):
        ws.cell(row=start_row, column=col).fill = FAILURE_DETAIL_FILL
        ws.cell(row=start_row, column=col).border = THIN_BORDER
    ws.merge_cells(f'A{start_row}:E{start_row}')
    message_cell = ws[f'A{start_row}']
    message_cell.value = f"Reason: {failure_details['message']}"
    message_cell.font = BOLD_FONT
    message_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    start_row += 1

    for col in range(1, 6):
        ws.cell(row=start_row, column=col).fill = FAILURE_DETAIL_FILL
        ws.cell(row=start_row, column=col).border = THIN_BORDER
    ws.merge_cells(f'A{start_row}:E{start_row}')
    suggestion_cell = ws[f'A{start_row}']
    suggestion_cell.value = f"Suggestion: {failure_details['suggestion']}"
    suggestion_cell.font = NORMAL_FONT
    suggestion_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    start_row += 2

    return start_row


# ============================================================
# EXCEL GENERATION FUNCTIONS
# ============================================================

def create_annual_balance_report_excel(financial_year: int, monthly_results: dict, output_folder: str) -> str:
    """
    Create annual Excel balance report with 12 monthly sheets.
    Each sheet contains the complete balance report for that month.
    
    Args:
        financial_year: Starting year of FY (e.g., 2025 for FY 2025-26)
        monthly_results: Dict with monthly calculation results (from run_full_year)
        output_folder: Folder to save the Excel file
    
    Returns:
        Path to the saved Excel file
    """
    print(f"GENERATING ANNUAL BALANCE SUMMARY EXCEL REPORT")
    print(f"{'='*70}")
    print(f"  Financial Year: FY {financial_year}-{str(financial_year + 1)[-2:]}")
    print(f"  Output Folder: {output_folder}")
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Define month order for financial year (April to March)
    fy_months = [
        (4, financial_year),      # April
        (5, financial_year),      # May
        (6, financial_year),      # June
        (7, financial_year),      # July
        (8, financial_year),      # August
        (9, financial_year),      # September
        (10, financial_year),     # October
        (11, financial_year),     # November
        (12, financial_year),     # December
        (1, financial_year + 1),  # January
        (2, financial_year + 1),  # February
        (3, financial_year + 1),  # March
    ]
    
    sheets_created = 0
    
    # Create a sheet for each month
    for month, year in fy_months:
        month_key = f"{year}_{month:02d}"
        month_data = monthly_results.get('months', {}).get(month_key)
        
        if not month_data:
            print(f"  ⚠ Skipping {MONTH_NAMES[month]} {year} - No data available")
            continue
        
        # Get full calculation result
        calculation_result = month_data.get('calculation_result')
        if not calculation_result:
            print(f"  ⚠ Skipping {MONTH_NAMES[month]} {year} - No calculation result stored")
            continue
        
        # Debug: Check what data is available
        has_usd_result = 'usd_result' in calculation_result
        has_final_dispatch = False
        if has_usd_result:
            usd_result = calculation_result.get('usd_result', {})
            has_final_dispatch = 'final_dispatch' in usd_result
        
        # Create sheet for this month (use short name to fit Excel limits)
        sheet_name = f"{MONTH_NAMES[month][:3]}-{year}"
        ws = wb.create_sheet(title=sheet_name)
        
        # Set initial column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        
        current_row = 1
        
        # Title
        for col in range(1, 6):
            ws.cell(row=current_row, column=col).fill = SECTION_FILL
        ws.merge_cells(f'A{current_row}:E{current_row}')
        title_cell = ws[f'A{current_row}']
        title_cell.value = f"Power & Utility Loop Balance Summary - {MONTH_NAMES[month]} {year}"
        title_cell.font = SECTION_FONT
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 2

        current_row = write_failure_banner(ws, current_row, calculation_result)
        
        # SECTION I: FUEL DEMAND
        current_row = write_fuel_demand_section(ws, current_row, month, year, calculation_result)
        current_row += 2
        
        # SECTION II: POWER BALANCE
        current_row = write_power_balance_section(ws, current_row, month, year, calculation_result)
        current_row += 2
        
        # SECTION III: STEAM BALANCE
        current_row = write_steam_balance_section(ws, current_row, month, year, calculation_result)
        current_row += 2
        
        # SECTION IV: OTHER UTILITIES BALANCE (chemicals are included inside this section)
        from services.other_utilities_balance_writer import write_other_utilities_balance_section
        current_row = write_other_utilities_balance_section(ws, current_row, month, year, calculation_result)
        current_row += 2
        
        # SECTION V: ASSET AVAILABILITY AND LOADING
        current_row = write_asset_availability_section(ws, current_row, month, year, calculation_result)

        # Auto-fit column widths based on content
        from openpyxl.cell.cell import MergedCell
        for column_cells in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column_cells:
                if isinstance(cell, MergedCell):
                    continue
                if column_letter is None:
                    column_letter = cell.column_letter
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = max(adjusted_width, 12)

        sheets_created += 1
    
    # If no sheets were created, create a summary sheet
    if sheets_created == 0:
        ws = wb.create_sheet(title="Summary")
        ws['A1'] = "No monthly data available for this financial year"
        ws['A1'].font = SECTION_FONT
        print(f"  ⚠ No sheets created - no successful monthly calculations found")
    
    # Save file
    os.makedirs(output_folder, exist_ok=True)
    filename = f"Annual_Balance_Summary_FY{financial_year}_{financial_year+1}.xlsx"
    filepath = os.path.join(output_folder, filename)
    
    wb.save(filepath)
    print(f"  [OK] Annual Excel report saved: {filepath}")
    print(f"  [OK] Total sheets created: {sheets_created}/12")
    print(f"{'='*70}\n")
    
    return filepath


def create_balance_report_excel(month: int, year: int, calculation_result: dict, output_folder: str) -> str:
    """
    Generate Power & Utility Loop Balance Summary Excel report for a single month.
    
    Args:
        month: Month number (1-12)
        year: Year
        calculation_result: Calculation result dictionary from calculate_budget_with_iteration
        output_folder: Folder to save the Excel file
    
    Returns:
        Path to the generated Excel file
    """
    print(f"\n{'='*70}")
    print(f"GENERATING BALANCE SUMMARY EXCEL REPORT")
    print(f"{'='*70}")
    print(f"  Month: {MONTH_NAMES[month]} {year}")
    print(f"  Output Folder: {output_folder}")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"{MONTH_NAMES[month]}-{year}"
    
    # Column widths will be auto-fitted after all content is written
    # Initial minimum widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    
    current_row = 1
    
    # Title
    for col in range(1, 6):
        ws.cell(row=current_row, column=col).fill = SECTION_FILL
    ws.merge_cells(f'A{current_row}:E{current_row}')
    title_cell = ws[f'A{current_row}']
    title_cell.value = f"Power & Utility Loop Balance Summary - {MONTH_NAMES[month]} {year}"
    title_cell.font = SECTION_FONT
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    current_row += 2

    current_row = write_failure_banner(ws, current_row, calculation_result)
    
    # SECTION I: FUEL DEMAND
    current_row = write_fuel_demand_section(ws, current_row, month, year, calculation_result)
    current_row += 2
    
    # SECTION II: POWER BALANCE
    current_row = write_power_balance_section(ws, current_row, month, year, calculation_result)
    current_row += 2
    
    # SECTION III: STEAM BALANCE
    current_row = write_steam_balance_section(ws, current_row, month, year, calculation_result)
    current_row += 2
    
    # SECTION IV: OTHER UTILITIES BALANCE
    from services.other_utilities_balance_writer import write_other_utilities_balance_section
    current_row = write_other_utilities_balance_section(ws, current_row, month, year, calculation_result)
    current_row += 2
    
    # SECTION V: ASSET AVAILABILITY AND LOADING
    current_row = write_asset_availability_section(ws, current_row, month, year, calculation_result)
    
    # Auto-fit column widths based on content
    from openpyxl.cell.cell import MergedCell
    
    for column_cells in ws.columns:
        max_length = 0
        column_letter = None
        
        for cell in column_cells:
            # Skip merged cells
            if isinstance(cell, MergedCell):
                continue
            
            # Get column letter from first non-merged cell
            if column_letter is None:
                column_letter = cell.column_letter
            
            try:
                if cell.value:
                    # Calculate length considering the cell value
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        # Set column width with some padding (add 2 for padding)
        if column_letter:
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 to prevent extremely wide columns
            ws.column_dimensions[column_letter].width = max(adjusted_width, 12)  # Minimum width of 12
    
    # Save file
    os.makedirs(output_folder, exist_ok=True)
    filename = f"Balance_Summary_{year}_{month:02d}_{MONTH_NAMES[month]}.xlsx"
    filepath = os.path.join(output_folder, filename)
    
    wb.save(filepath)
    print(f"  [OK] Excel report saved: {filepath}")
    print(f"{'='*70}\n")
    
    return filepath


def write_fuel_demand_section(ws, start_row: int, month: int, year: int, calculation_result: dict) -> int:
    """Write Section I: Fuel Demand to worksheet - Single unified table for all assets."""
    fuel_data = extract_fuel_demand_data(month, year, calculation_result)
    
    row = start_row
    
    # Section header
    for col in range(1, 7):
        ws.cell(row=row, column=col).fill = SECTION_FILL
    ws.merge_cells(f'A{row}:F{row}')
    header_cell = ws[f'A{row}']
    header_cell.value = "SECTION I: FUEL DEMAND"
    header_cell.font = SECTION_FONT
    header_cell.alignment = Alignment(horizontal='left', vertical='center')
    row += 1
    
    # Single table header for all assets
    headers = ['Asset', 'Asset Type', 'NCV', 'Quantity (MMBTU)', 'Allocated Load', 'Heat Rate']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = BOLD_FONT
        cell.fill = SUBSECTION_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Set column widths for better formatting
    ws.column_dimensions['A'].width = 12  # Asset
    ws.column_dimensions['B'].width = 15  # Asset Type
    ws.column_dimensions['C'].width = 18  # NCV
    ws.column_dimensions['D'].width = 18  # Quantity (MMBTU)
    ws.column_dimensions['E'].width = 18  # Allocated Load
    ws.column_dimensions['F'].width = 18  # Heat Rate
    
    row += 1
    
    # GT Assets
    for gt in fuel_data['gt_assets']:
        ws[f'A{row}'] = gt['asset_name']
        ws[f'B{row}'] = 'Gas Turbine'
        ws[f'C{row}'] = f"{round(gt['ncv_kcal_kwh'], 2)} GBT/MT"
        ws[f'D{row}'] = round(gt['quantity_mmbtu'], 2)
        # GT allocated load shows average MW per hour
        ws[f'E{row}'] = f"{round(gt['allocated_load_mw'], 2)} MW"
        ws[f'F{row}'] = f"{round(gt['heat_rate_kcal_kwh'], 2)} Kcal/kWh"
        
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1
    
    # HRSG Assets
    for hrsg in fuel_data['hrsg_assets']:
        ws[f'A{row}'] = hrsg['asset_name']
        ws[f'B{row}'] = 'HRSG'
        ws[f'C{row}'] = f"{round(hrsg['ncv_gbt'], 2)} GBT/MT"  # Use NCV from database (GBT/MT)
        ws[f'D{row}'] = round(hrsg['quantity_mmbtu'], 2)
        ws[f'E{row}'] = f"{round(hrsg['allocated_load_mt_per_hr'], 2)} MT/h"  # Average hourly load
        ws[f'F{row}'] = f"{round(hrsg['heat_rate_btu_lb'], 2)} Kcal/kg"
        
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1
    
    return row


def write_power_balance_section(ws, start_row: int, month: int, year: int, calculation_result: dict) -> int:
    """Write Section II: Power Balance with 5-column layout (Demand, Norm, Quantity MW, Generation, Quantity MW)."""
    power_data = extract_power_balance_data(month, year, calculation_result)
    usd_result = calculation_result.get('usd_result', {})
    final_dispatch = usd_result.get('final_dispatch', [])
    utility_consumption = calculation_result.get('utility_consumption', {})
    
    # Month-wise norms from DB
    norm_gt_aux = {
        "GT1": get_utility_norm_from_db(
            month, year, "NMD - Power Plant 1", "POWERGEN", "Power_Dis",
            issuing_plant_name="NMD - Utility/Power Dist",
        ),
        "GT2": get_utility_norm_from_db(
            month, year, "NMD - Power Plant 2", "POWERGEN", "Power_Dis",
            issuing_plant_name="NMD - Utility/Power Dist",
        ),
        "GT3": get_utility_norm_from_db(
            month, year, "NMD - Power Plant 3", "POWERGEN", "Power_Dis",
            issuing_plant_name="NMD - Utility/Power Dist",
        ),
    }
    norm_stg_aux = get_utility_norm_from_db(
        month, year, "NMD - STG Power Plant", "POWERGEN", "Power_Dis",
        issuing_plant_name="NMD - Utility/Power Dist",
    )
    utility_power_norms = {
        "BFW Plant": get_utility_norm_from_db(month, year, "NMD - Utility Plant", "Boiler Feed Water", "Power_Dis"),
        "DM Water Plant": get_utility_norm_from_db(month, year, "NMD - Utility Plant", "D M Water", "Power_Dis"),
        "Cooling Water 1": get_utility_norm_from_db(month, year, "NMD - Utility Plant", "Cooling Water 1", "Power_Dis"),
        "Cooling Water 2": get_utility_norm_from_db(month, year, "NMD - Utility Plant", "Cooling Water 2", "Power_Dis"),
        "Compressed Air": get_utility_norm_from_db(month, year, "NMD - Utility Plant", "COMPRESSED AIR", "Power_Dis"),
        "Oxygen Plant": get_utility_norm_from_db(month, year, "NMD - Utility Plant", "Oxygen", "Power_Dis"),
        "Effluent Treatment": get_utility_norm_from_db(month, year, "NMD - Utility Plant", "Effluent Treated", "Power_Dis"),
    }
    
    row = start_row
    
    # Section header
    for col in range(1, 6):
        ws.cell(row=row, column=col).fill = SECTION_FILL
    ws.merge_cells(f'A{row}:E{row}')
    header_cell = ws[f'A{row}']
    header_cell.value = "SECTION II: POWER BALANCE"
    header_cell.font = SECTION_FONT
    header_cell.alignment = Alignment(horizontal='left', vertical='center')
    row += 1
    
    # Column headers
    headers = ['Demand', 'Norm', 'Quantity MW', 'Generation by Asset', 'Quantity MW']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = BOLD_FONT
        cell.fill = SUBSECTION_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    start_data_row = row
    
    # LEFT SIDE: DEMAND
    # Plant Requirement
    ws[f'A{row}'] = "Plant Requirement (Power)"
    ws[f'A{row}'].font = BOLD_FONT
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = THIN_BORDER
    row += 1
    
    plant_req_total = 0
    for plant_data in power_data['demand']['plant_requirement']:
        ws[f'A{row}'] = f"  {plant_data['plant_name']}"
        demand_mw = plant_data['demand_value'] / 1000  # Convert KWH to MWH
        ws[f'C{row}'] = round(demand_mw, 2)
        plant_req_total += demand_mw
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1
    
    # Fixed Requirement
    ws[f'A{row}'] = "Fixed Requirement (Power)"
    ws[f'A{row}'].font = BOLD_FONT
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = THIN_BORDER
    row += 1
    
    fixed_req_total = 0
    for fixed_data in power_data['demand']['fixed_requirement']:
        ws[f'A{row}'] = f"  {fixed_data['consumer_name']} ({fixed_data['plant_name']})"
        demand_mw = fixed_data['consumption_value'] / 1000  # Convert KWH to MWH
        ws[f'C{row}'] = round(demand_mw, 2)
        fixed_req_total += demand_mw
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1
    
    # Utility for Utility (U4U)
    ws[f'A{row}'] = "Utility for Utility (U4U)"
    ws[f'A{row}'].font = BOLD_FONT
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = THIN_BORDER
    row += 1
    
    aux_power_total = 0
    # Extract GT and STG generation and calculate auxiliary power
    gt_generation = {'GT1': 0, 'GT2': 0, 'GT3': 0}
    stg_generation = 0
    
    for asset in final_dispatch:
        asset_name = asset.get('AssetName', '')
        gross_mwh = asset.get('GrossMWh', 0)
        
        if 'Plant-1' in asset_name or 'Plant 1' in asset_name:
            gt_generation['GT1'] = gross_mwh
        elif 'Plant-2' in asset_name or 'Plant 2' in asset_name:
            gt_generation['GT2'] = gross_mwh
        elif 'Plant-3' in asset_name or 'Plant 3' in asset_name:
            gt_generation['GT3'] = gross_mwh
        elif 'STG' in asset_name:
            stg_generation = gross_mwh
    
    # GT auxiliary power
    for gt_name in ['GT1', 'GT2', 'GT3']:
        ws[f'A{row}'] = f"  {gt_name} - Power"
        gt_aux_norm = norm_gt_aux.get(gt_name, 0)
        ws[f'B{row}'] = gt_aux_norm
        aux_mw = gt_generation[gt_name] * gt_aux_norm
        ws[f'C{row}'] = round(aux_mw, 2)
        aux_power_total += aux_mw
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1
    
    # STG auxiliary power
    ws[f'A{row}'] = f"  STG - Power"
    ws[f'B{row}'] = norm_stg_aux
    stg_aux_mw = stg_generation * norm_stg_aux
    ws[f'C{row}'] = round(stg_aux_mw, 2)
    aux_power_total += stg_aux_mw
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = THIN_BORDER
    row += 1
    
    # Utility Plant Power (U4U - Utility for Utility)
    # Extract from final_u4u_power (matches console output) instead of utility_consumption
    # This ensures Excel shows the exact same values as the console log
    final_u4u_power = usd_result.get('final_u4u_power', {})
    utility_power_from_u4u = final_u4u_power.get('utility_power', {})
    
    # If final_u4u_power is not available (backward compatibility), fall back to utility_consumption
    if not utility_power_from_u4u:
        utility_power_from_u4u = utility_consumption.get('utility_power', {})
    
    utility_power_items = [
        ('BFW Plant', utility_power_from_u4u.get('bfw_kwh', 0) / 1000),
        ('DM Water Plant', utility_power_from_u4u.get('dm_kwh', 0) / 1000),
        ('Cooling Water 1', utility_power_from_u4u.get('cw1_kwh', 0) / 1000),
        ('Cooling Water 2', utility_power_from_u4u.get('cw2_kwh', 0) / 1000),
        ('Compressed Air', utility_power_from_u4u.get('air_kwh', 0) / 1000),
        ('Oxygen Plant', utility_power_from_u4u.get('oxygen_kwh', 0) / 1000),
        ('Effluent Treatment', utility_power_from_u4u.get('effluent_kwh', 0) / 1000),
    ]
    
    for utility_name, power_mwh in utility_power_items:
        norm_value = utility_power_norms.get(utility_name, 0)
        ws[f'A{row}'] = f"  {utility_name}"
        ws[f'B{row}'] = norm_value  # Add norm value
        ws[f'C{row}'] = round(power_mwh, 2)
        aux_power_total += power_mwh
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1
    
    # Get actual total from calculation result
    usd_power = usd_result.get('power_result', {})
    total_demand_actual = usd_power.get('totalDemandUnits', 0)
    
    # Since we're now using final_u4u_power values, aux_power_total matches the engine's u4uPower
    # No reconciliation needed - the values are identical
    line_items_sum = plant_req_total + fixed_req_total + aux_power_total
    
    # RIGHT SIDE: GENERATION (start from top)
    gen_row = start_data_row
    
    # Generation assets — always show all four assets + Import Power, even if 0.
    # Build lookup from dispatched assets.
    gen_map = {'GT1': 0.0, 'GT2': 0.0, 'GT3': 0.0, 'STG': 0.0}
    for asset in final_dispatch:
        asset_name = asset.get('AssetName', '')
        net_mwh = asset.get('NetMWh', 0)
        if 'Plant-1' in asset_name or 'Plant 1' in asset_name:
            gen_map['GT1'] = net_mwh
        elif 'Plant-2' in asset_name or 'Plant 2' in asset_name:
            gen_map['GT2'] = net_mwh
        elif 'Plant-3' in asset_name or 'Plant 3' in asset_name:
            gen_map['GT3'] = net_mwh
        elif 'STG' in asset_name:
            gen_map['STG'] = net_mwh
    
    import_mwh = usd_power.get('mandatoryImportUsed', 0)
    
    generation_total = 0
    for asset_label in ['GT1', 'GT2', 'GT3', 'STG']:
        val = gen_map[asset_label]
        ws[f'D{gen_row}'] = asset_label
        ws[f'E{gen_row}'] = round(val, 2)
        generation_total += val
        for col in range(1, 6):
            ws.cell(row=gen_row, column=col).border = THIN_BORDER
        gen_row += 1
    
    # Import Power always shown
    ws[f'D{gen_row}'] = 'Import Power'
    ws[f'E{gen_row}'] = round(import_mwh, 2)
    generation_total += import_mwh
    for col in range(1, 6):
        ws.cell(row=gen_row, column=col).border = THIN_BORDER
    gen_row += 1
    
    # Summary rows at bottom (no blank row - keep totals adjacent to data)
    row = max(row, gen_row)
    
    # Use actual total from calculation result
    total_demand = total_demand_actual
    
    # Total row
    ws[f'A{row}'] = "Total Power Demand"
    ws[f'C{row}'] = round(total_demand, 2)
    ws[f'D{row}'] = "Total Power Generation"
    ws[f'E{row}'] = round(generation_total, 2)
    for col in [1, 2, 3, 4, 5]:
        ws.cell(row=row, column=col).font = BOLD_FONT
        ws.cell(row=row, column=col).fill = TOTAL_FILL
        ws.cell(row=row, column=col).border = THIN_BORDER
    row += 1
    
    # Imbalance row
    imbalance = generation_total - total_demand
    ws[f'A{row}'] = "Power Imbalance"
    ws[f'C{row}'] = round(imbalance, 2)
    for col in range(1, 6):
        ws.cell(row=row, column=col).font = BOLD_FONT
        ws.cell(row=row, column=col).fill = TOTAL_FILL
        ws.cell(row=row, column=col).border = THIN_BORDER
    row += 1
    
    return row + 1


def write_steam_balance_section(ws, start_row: int, month: int, year: int, calculation_result: dict) -> int:
    """Write Section III: Steam Balance with 5-column layout for LP, MP, HP, SHP."""
    row = start_row
    
    # Section header
    for col in range(1, 6):
        ws.cell(row=row, column=col).fill = SECTION_FILL
    ws.merge_cells(f'A{row}:E{row}')
    header_cell = ws[f'A{row}']
    header_cell.value = "SECTION III: STEAM BALANCE"
    header_cell.font = SECTION_FONT
    header_cell.alignment = Alignment(horizontal='left', vertical='center')
    row += 1
    
    # Write balance for each steam type in order: LP → MP → HP → SHP
    for steam_type in ['LP', 'MP', 'HP', 'SHP']:
        row = write_single_steam_balance(ws, row, month, year, calculation_result, steam_type)
        row += 1  # Add spacing between steam types
    
    return row


def write_single_steam_balance(ws, start_row: int, month: int, year: int, calculation_result: dict, steam_type: str) -> int:
    """Write a single steam type balance with 5-column layout (Demand, Norm, Quantity MT, Generation, Quantity MT)."""
    usd_result = calculation_result.get('usd_result', {})
    final_steam = usd_result.get('final_steam_balance', {})
    stg_extraction = calculation_result.get('stg_extraction', {})
    utility_consumption = calculation_result.get('utility_consumption', {})
    hrsg_full_load_mode = calculation_result.get('hrsg_full_load_mode', False)
    
    steam_type_lower = steam_type.lower()
    steam_balance = final_steam.get(f'{steam_type_lower}_balance', {})
    
    # Map steam type to utility name in database
    utility_name_map = {
        'SHP': 'SHP Steam_Dis',
        'HP': 'HP Steam_Dis',
        'MP': 'MP Steam_Dis',
        'LP': 'LP Steam_Dis'
    }
    utility_name = utility_name_map.get(steam_type, f'{steam_type} Steam_Dis')
    
    row = start_row
    
    # Subsection header
    for col in range(1, 6):
        ws.cell(row=row, column=col).fill = SUBSECTION_FILL
    ws.merge_cells(f'A{row}:E{row}')
    header_cell = ws[f'A{row}']
    header_cell.value = f"{steam_type} STEAM BALANCE"
    header_cell.font = Font(bold=True, size=11)
    header_cell.alignment = Alignment(horizontal='left', vertical='center')
    row += 1
    
    # Column headers
    headers = ['Demand', 'Norm', 'Quantity MT', 'Generation by Asset', 'Quantity MT']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = BOLD_FONT
        cell.fill = SUBSECTION_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    start_data_row = row
    
    # LEFT SIDE: DEMAND
    # Plant Requirement
    ws[f'A{row}'] = f"Plant Requirement ({steam_type} Steam)"
    ws[f'A{row}'].font = BOLD_FONT
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = THIN_BORDER
    row += 1
    
    plant_req_total = 0
    plant_requirements = get_plant_wise_demand(month, year, utility_name)
    for plant_data in plant_requirements:
        ws[f'A{row}'] = f"  {plant_data['plant_name']}"
        # Steam is already in MT in database, no conversion needed
        # Only Power (KWH) needs conversion to MWH
        demand_mt = plant_data['demand_value']
        ws[f'C{row}'] = round(demand_mt, 2)
        plant_req_total += demand_mt
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1
    
    # Fixed Requirement
    ws[f'A{row}'] = f"Fixed Requirement ({steam_type} Steam)"
    ws[f'A{row}'].font = BOLD_FONT
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = THIN_BORDER
    row += 1
    
    fixed_req_total = 0
    fixed_requirements = get_fixed_consumption_details(month, year, utility_name)
    for fixed_data in fixed_requirements:
        ws[f'A{row}'] = f"  {fixed_data['consumer_name']} ({fixed_data['plant_name']})"
        # Steam is already in MT in database, no conversion needed
        consumption_mt = fixed_data['consumption_value']
        ws[f'C{row}'] = round(consumption_mt, 2)
        fixed_req_total += consumption_mt
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1
    
    # Utility for Utility (U4U)
    ws[f'A{row}'] = "Utility for Utility (U4U)"
    ws[f'A{row}'].font = BOLD_FONT
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = THIN_BORDER
    row += 1
    
    u4u_total = 0
    
    # Add steam-specific U4U items based on steam type
    if steam_type == 'SHP':
        # Get STG extraction lookup data for display
        stg_ext = stg_extraction if stg_extraction else {}
        stg_op_hrs = stg_ext.get('stg_operating_hours', 0)
        sp_steam_power_norm = stg_ext.get('sp_steam_power', 0)  # MT/MWH from lookup
        
        # STG Power Generation = SpSteamPower × Gross MWH (= SVHInletTPH × Hours)
        stg_shp_power = steam_balance.get('stg_shp_power', 0)
        # Get STG gross MWH for norm display
        stg_gross_kwh = stg_ext.get('stg_gross_kwh', 0)
        stg_gross_mwh = stg_gross_kwh / 1000 if stg_gross_kwh > 0 else 0
        
        ws[f'A{row}'] = "  STG Power Generation"
        if sp_steam_power_norm > 0:
            ws[f'B{row}'] = round(sp_steam_power_norm, 4)  # SpSteamPower norm (MT/MWH)
        ws[f'C{row}'] = round(stg_shp_power, 2)
        u4u_total += stg_shp_power
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1
        
        # --- Informational breakdown of STG inlet steam ---
        # LP Extraction from STG = SLExtFlowTPH × Operating Hours (from lookup table)
        lp_ext_tph = stg_ext.get('lp_extraction_tph', 0)
        lp_ext_mt = lp_ext_tph * stg_op_hrs
        ws[f'A{row}'] = "  LP Extraction (STG)"
        ws[f'C{row}'] = round(lp_ext_mt, 2)
        # NOT added to u4u_total - already part of STG Power Gen (SVHInletTPH)
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1
        
        # MP Extraction from STG = SMBleedFlowTPH × Operating Hours (from lookup table)
        mp_ext_tph = stg_ext.get('mp_extraction_tph', 0)
        mp_ext_mt = mp_ext_tph * stg_op_hrs
        ws[f'A{row}'] = "  MP Extraction (STG)"
        ws[f'C{row}'] = round(mp_ext_mt, 2)
        # NOT added to u4u_total - already part of STG Power Gen (SVHInletTPH)
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

        
        # SHP for MP via PRDS
        lp_balance = final_steam.get('lp_balance', {})
        mp_balance = final_steam.get('mp_balance', {})
        shp_for_mp_prds = mp_balance.get('shp_for_prds_mp', 0)
        ws[f'A{row}'] = "  MP Steam (via PRDS)"
        ws[f'C{row}'] = round(shp_for_mp_prds, 2)
        u4u_total += shp_for_mp_prds
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1
        
        # SHP for HP via PRDS
        shp_for_hp_prds = steam_balance.get('shp_for_hp_prds', 0)
        ws[f'A{row}'] = "  HP Steam (via PRDS)"
        ws[f'C{row}'] = round(shp_for_hp_prds, 2)
        u4u_total += shp_for_hp_prds
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1
        
        if hrsg_full_load_mode:
            # Free steam is considered negative demand
            hrsg_dispatch = usd_result.get('hrsg_dispatch', {})
            hrsg_dispatch_list = hrsg_dispatch.get('hrsg_dispatch', [])
            
            for h in hrsg_dispatch_list:
                free_mt = h.get('free_steam_mt', 0)
                if free_mt > 0:
                    ws[f'A{row}'] = f"  {h.get('linked_gt', 'GT')} (Free Steam Offset)"
                    ws[f'C{row}'] = -round(free_mt, 2)
                    u4u_total -= free_mt
                    for col in range(1, 6):
                        ws.cell(row=row, column=col).border = THIN_BORDER
                    row += 1
    
    elif steam_type == 'MP':
        # MP consumed by PRDS to generate LP (steam cascade)
        mp_for_lp = steam_balance.get('mp_for_lp', 0)
        lp_from_prds_mt = usd_result.get('final_steam_balance', {}).get('lp_balance', {}).get('lp_from_prds', 0.0)
        lp_prds_norm_db = get_utility_norm_from_db(month, year, 'NMD - Utility Plant', 'LP Steam PRDS', 'MP Steam_Dis')
        lp_prds_norm = round(mp_for_lp / lp_from_prds_mt, 4) if lp_from_prds_mt else lp_prds_norm_db
        ws[f'A{row}'] = "  LP Steam (via PRDS)"
        ws[f'B{row}'] = lp_prds_norm
        ws[f'C{row}'] = round(mp_for_lp, 2)
        u4u_total += mp_for_lp
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

        # Treated Spent Caustic (TSC) — genuine MP U4U (separate from steam cascade)
        final_u4u_mp = usd_result.get('final_u4u_mp_steam', {}) or {}
        tsc_norm = final_u4u_mp.get('tsc_norm', 0)
        tsc_mt = final_u4u_mp.get('tsc_mt', 0)
        ws[f'A{row}'] = "  Treated Spent Caustic (U4U)"
        ws[f'B{row}'] = round(tsc_norm, 6) if tsc_norm else ''
        ws[f'C{row}'] = round(tsc_mt, 2)
        u4u_total += tsc_mt
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    elif steam_type == 'HP':
        pass

    elif steam_type == 'LP':
        # LP consumed by BFW deaerator and HRSG drum heating
        final_u4u_lp = usd_result.get('final_u4u_lp_steam', {}) or {}
        hrsg_dispatch = usd_result.get('hrsg_dispatch', {}) or {}
        hrsg_dispatch_list = hrsg_dispatch.get('hrsg_dispatch', [])
        hrsg_shp_map = {'HRSG1': 0.0, 'HRSG2': 0.0, 'HRSG3': 0.0}
        for hrsg_data in hrsg_dispatch_list:
            hrsg_name = (hrsg_data.get('name', '') or '').upper().replace('-', '').replace(' ', '')
            total_shp_mt = hrsg_data.get('total_shp_mt', 0) or 0
            if 'HRSG1' in hrsg_name:
                hrsg_shp_map['HRSG1'] = total_shp_mt
            elif 'HRSG2' in hrsg_name:
                hrsg_shp_map['HRSG2'] = total_shp_mt
            elif 'HRSG3' in hrsg_name:
                hrsg_shp_map['HRSG3'] = total_shp_mt

        bfw = utility_consumption.get('bfw', {})
        bfw_total_m3 = bfw.get('total_m3', 0) or 0
        bfw_norm = final_u4u_lp.get('bfw_norm', 0) or get_utility_norm_from_db(month, year, 'NMD - Utility Plant', 'Boiler Feed Water', 'LP Steam_Dis')
        hrsg1_norm = final_u4u_lp.get('hrsg1_norm', 0) or get_utility_norm_from_db(month, year, 'NMD - Utility Plant', 'HRSG1_SHP STEAM', 'LP Steam_Dis')
        hrsg2_norm = final_u4u_lp.get('hrsg2_norm', 0) or get_utility_norm_from_db(month, year, 'NMD - Utility Plant', 'HRSG2_SHP STEAM', 'LP Steam_Dis')
        hrsg3_norm = final_u4u_lp.get('hrsg3_norm', 0) or get_utility_norm_from_db(month, year, 'NMD - Utility Plant', 'HRSG3_SHP STEAM', 'LP Steam_Dis')

        lp_u4u_items = [
            ('BFW Plant', final_u4u_lp.get('bfw_mt', 0) or (bfw_total_m3 * bfw_norm), bfw_norm),
            ('HRSG1', final_u4u_lp.get('hrsg1_mt', 0) or (hrsg_shp_map['HRSG1'] * hrsg1_norm), hrsg1_norm),
            ('HRSG2', final_u4u_lp.get('hrsg2_mt', 0) or (hrsg_shp_map['HRSG2'] * hrsg2_norm), hrsg2_norm),
            ('HRSG3', final_u4u_lp.get('hrsg3_mt', 0) or (hrsg_shp_map['HRSG3'] * hrsg3_norm), hrsg3_norm),
        ]
        for item_name, item_mt, item_norm in lp_u4u_items:
            ws[f'A{row}'] = f"  {item_name}"
            ws[f'B{row}'] = round(item_norm, 6) if item_norm else ''
            ws[f'C{row}'] = round(item_mt, 2)
            u4u_total += item_mt
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = THIN_BORDER
            row += 1
    
    # Calculate adjustment between line items and actual calculation total
    usd_steam = usd_result.get('final_steam_balance', {})
    steam_result = usd_steam.get(f'{steam_type_lower}_balance', {})
    
    # SHP uses 'shp_total_demand' key, others use '{type}_total'
    if steam_type == 'SHP':
        total_demand_actual = steam_result.get('shp_total_demand', 0)
        if hrsg_full_load_mode:
            hrsg_dispatch = usd_result.get('hrsg_dispatch', {})
            total_free = sum(h.get('free_steam_mt', 0) for h in hrsg_dispatch.get('hrsg_dispatch', []))
            total_demand_actual -= total_free
    else:
        total_demand_actual = steam_result.get(f'{steam_type_lower}_total', 0)
    
    line_items_sum = plant_req_total + fixed_req_total + u4u_total
    displayed_total_demand = line_items_sum
    
    # Verify the line items sum matches the calculation total.
    # For LP steam, the engine's lp_total = lp_process + lp_fixed + lp_ufu (converged U4U).
    # The displayed sum = plant_req_total + fixed_req_total + u4u_total.
    # These should be very close after LP U4U convergence is enforced in the iteration loop.
    # A small residual difference (<1 MT) is acceptable due to rounding between iterations.
    if abs(total_demand_actual - line_items_sum) > 1.0:
        engine_ufu = steam_result.get(f'{steam_type_lower}_ufu', 0) if steam_type != 'SHP' else 0
        print(f"[WARNING] {steam_type} Steam Balance mismatch:")
        print(f"  Displayed demand sum: {line_items_sum:.2f} MT  (plant={plant_req_total:.2f} + fixed={fixed_req_total:.2f} + u4u={u4u_total:.2f})")
        print(f"  Engine total:         {total_demand_actual:.2f} MT  (engine u4u={engine_ufu:.2f})")
        print(f"  Difference:           {total_demand_actual - line_items_sum:.2f} MT")
    
    # RIGHT SIDE: GENERATION (start from top)
    gen_row = start_data_row
    
    # Generation sources
    generation_total = 0
    
    # Extract supply sources based on steam type
    if steam_type == 'SHP':
        hrsg_dispatch = usd_result.get('hrsg_dispatch', {})
        hrsg_dispatch_list = hrsg_dispatch.get('hrsg_dispatch', [])

        # SHP generated (Fired + Free Steam)
        hrsg_shp_gen = {
            'HRSG-1 (Supp Firing)': 0.0, 'GT-1 (Free Steam)': 0.0,
            'HRSG-2 (Supp Firing)': 0.0, 'GT-2 (Free Steam)': 0.0,
            'HRSG-3 (Supp Firing)': 0.0, 'GT-3 (Free Steam)': 0.0
        }
        for hrsg_data in hrsg_dispatch_list:
            raw_name = hrsg_data.get('name', '')
            dispatched_supp_mt = hrsg_data.get('dispatched_supp_mt', 0) or 0
            free_steam_mt = hrsg_data.get('free_steam_mt', 0) or 0
            
            norm = (raw_name.upper()
                    .replace('NMD-', '').replace('NMD ', '')
                    .replace('_SHP STEAM', '').replace('SHP STEAM', '')
                    .strip())
            # Map to canonical labels
            if '1' in norm:
                hrsg_shp_gen['HRSG-1 (Supp Firing)'] = dispatched_supp_mt
                if not hrsg_full_load_mode:
                    hrsg_shp_gen['GT-1 (Free Steam)'] = free_steam_mt
            elif '2' in norm:
                hrsg_shp_gen['HRSG-2 (Supp Firing)'] = dispatched_supp_mt
                if not hrsg_full_load_mode:
                    hrsg_shp_gen['GT-2 (Free Steam)'] = free_steam_mt
            elif '3' in norm:
                hrsg_shp_gen['HRSG-3 (Supp Firing)'] = dispatched_supp_mt
                if not hrsg_full_load_mode:
                    hrsg_shp_gen['GT-3 (Free Steam)'] = free_steam_mt
        
        # Add to excel
        for hrsg_label, shp_mt in hrsg_shp_gen.items():
            if shp_mt > 0 or 'Supp Firing' in hrsg_label: # Always show HRSG rows even if 0, but only show GT rows if >0
                ws[f'D{gen_row}'] = hrsg_label
                ws[f'E{gen_row}'] = round(shp_mt, 2)
                generation_total += shp_mt
                for col in range(1, 6):
                    ws.cell(row=gen_row, column=col).border = THIN_BORDER
                gen_row += 1
    
    elif steam_type == 'HP':
        # HP from PRDS (SHP → HP) — always show
        hp_from_prds = steam_balance.get('hp_from_prds', 0)
        ws[f'D{gen_row}'] = "HP PRDS (from SHP)"
        ws[f'E{gen_row}'] = round(hp_from_prds, 2)
        generation_total += hp_from_prds
        for col in range(1, 6):
            ws.cell(row=gen_row, column=col).border = THIN_BORDER
        gen_row += 1
    
    elif steam_type == 'MP':
        # MP from PRDS (SHP → MP) — always show
        mp_from_prds = steam_balance.get('mp_from_prds', 0)
        ws[f'D{gen_row}'] = "MP PRDS (from SHP)"
        ws[f'E{gen_row}'] = round(mp_from_prds, 2)
        generation_total += mp_from_prds
        for col in range(1, 6):
            ws.cell(row=gen_row, column=col).border = THIN_BORDER
        gen_row += 1
        
        # MP from STG extraction — always show
        mp_from_stg = steam_balance.get('mp_from_stg', 0)
        ws[f'D{gen_row}'] = "STG Extraction"
        ws[f'E{gen_row}'] = round(mp_from_stg, 2)
        generation_total += mp_from_stg
        for col in range(1, 6):
            ws.cell(row=gen_row, column=col).border = THIN_BORDER
        gen_row += 1
    
    elif steam_type == 'LP':
        # LP from PRDS (MP → LP) — always show
        lp_from_prds = steam_balance.get('lp_from_prds', 0)
        ws[f'D{gen_row}'] = "LP PRDS (from MP)"
        ws[f'E{gen_row}'] = round(lp_from_prds, 2)
        generation_total += lp_from_prds
        for col in range(1, 6):
            ws.cell(row=gen_row, column=col).border = THIN_BORDER
        gen_row += 1
        
        # LP from STG extraction — always show
        lp_from_stg = steam_balance.get('lp_from_stg', 0)
        ws[f'D{gen_row}'] = "STG Extraction"
        ws[f'E{gen_row}'] = round(lp_from_stg, 2)
        generation_total += lp_from_stg
        for col in range(1, 6):
            ws.cell(row=gen_row, column=col).border = THIN_BORDER
        gen_row += 1
    
    # Summary rows at bottom (no blank row - keep totals adjacent to data)
    row = max(row, gen_row)
    
    # Excel total demand must match the sum of the displayed demand rows
    total_demand = displayed_total_demand
    
    # Total row
    ws[f'A{row}'] = f"Total {steam_type} Steam Demand"
    ws[f'C{row}'] = round(total_demand, 2)
    ws[f'D{row}'] = f"Total {steam_type} Steam Generation"
    ws[f'E{row}'] = round(generation_total, 2)
    for col in [1, 2, 3, 4, 5]:
        ws.cell(row=row, column=col).font = BOLD_FONT
        ws.cell(row=row, column=col).fill = TOTAL_FILL
        ws.cell(row=row, column=col).border = THIN_BORDER
    row += 1
    
    # Imbalance row
    imbalance = generation_total - total_demand
    ws[f'A{row}'] = f"{steam_type} Steam Imbalance"
    ws[f'C{row}'] = round(imbalance, 2)
    for col in range(1, 6):
        ws.cell(row=row, column=col).font = BOLD_FONT
        ws.cell(row=row, column=col).fill = TOTAL_FILL
        ws.cell(row=row, column=col).border = THIN_BORDER
    row += 1
    
    return row


def write_asset_availability_section(ws, start_row: int, month: int, year: int, calculation_result: dict) -> int:
    """Write Section V: Asset Availability and Loading to worksheet."""
    asset_data = extract_asset_availability_data(month, year, calculation_result)
    
    row = start_row
    
    # Section header
    for col in range(1, 6):
        ws.cell(row=row, column=col).fill = SECTION_FILL
    ws.merge_cells(f'A{row}:E{row}')
    header_cell = ws[f'A{row}']
    header_cell.value = "SECTION V: ASSET AVAILABILITY AND LOADING"
    header_cell.font = SECTION_FONT
    header_cell.alignment = Alignment(horizontal='left', vertical='center')
    row += 1
    
    # Generation Assets subsection
    for col in range(1, 6):
        ws.cell(row=row, column=col).fill = SUBSECTION_FILL
    ws.merge_cells(f'A{row}:E{row}')
    subheader_cell = ws[f'A{row}']
    subheader_cell.value = "Generation Assets"
    subheader_cell.font = Font(bold=True, size=11)
    subheader_cell.alignment = Alignment(horizontal='left', vertical='center')
    row += 1
    
    # Column headers for generation assets
    gen_headers = ['Asset', 'Availability, Hr', 'Min', 'Max', 'Plan / Allocation by loop']
    for col_idx, header in enumerate(gen_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = BOLD_FONT
        cell.fill = SUBSECTION_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    # Write generation assets data
    for asset in asset_data['generation_assets']:
        ws[f'A{row}'] = asset['asset_name']
        ws[f'B{row}'] = asset['availability_hr']
        ws[f'C{row}'] = asset['min_capacity']
        ws[f'D{row}'] = asset['max_capacity']
        ws[f'E{row}'] = asset['avg_load_per_hr']
        
        for col in [1, 2, 3, 4, 5]:
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
        
        row += 1
    
    # Add spacing
    row += 1
    
    # Steam Assets subsection
    for col in range(1, 6):
        ws.cell(row=row, column=col).fill = SUBSECTION_FILL
    ws.merge_cells(f'A{row}:E{row}')
    subheader_cell = ws[f'A{row}']
    subheader_cell.value = "Steam Assets"
    subheader_cell.font = Font(bold=True, size=11)
    subheader_cell.alignment = Alignment(horizontal='left', vertical='center')
    row += 1
    
    # Column headers for steam assets
    steam_headers = ['Asset', 'Availability, Hr', 'Min', 'Max', 'Plan / Allocation by loop']
    for col_idx, header in enumerate(steam_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = BOLD_FONT
        cell.fill = SUBSECTION_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    # Write steam assets data
    for asset in asset_data['steam_assets']:
        ws[f'A{row}'] = asset['asset_name']
        ws[f'B{row}'] = asset['availability_hr']
        ws[f'C{row}'] = asset['min_capacity']
        ws[f'D{row}'] = asset['max_capacity']
        ws[f'E{row}'] = asset['avg_load_per_hr']
        
        for col in [1, 2, 3, 4, 5]:
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
        
        row += 1
    
    return row


def write_other_utilities_section(ws, start_row: int, month: int, year: int, calculation_result: dict) -> int:
    """Write Section IV+: Other Utilities Balance to worksheet."""
    row = start_row
    
    # Section header
    for col in range(1, 4):
        ws.cell(row=row, column=col).fill = SECTION_FILL
    ws.merge_cells(f'A{row}:C{row}')
    header_cell = ws[f'A{row}']
    header_cell.value = "SECTION IV+: OTHER UTILITIES BALANCE"
    header_cell.font = SECTION_FONT
    header_cell.alignment = Alignment(horizontal='left', vertical='center')
    row += 1
    
    # Define utilities to include
    utilities = [
        ('Boiler Feed Water', 'bfw', 'M3'),
        ('D M Water', 'dm_water', 'M3'),
        ('Cooling Water 1', 'cooling_water', 'KM3'),
        ('Cooling Water 2', 'cooling_water', 'KM3'),
        ('COMPRESSED AIR', 'compressed_air', 'NM3'),
        ('Oxygen', 'oxygen', 'MT'),
        ('Effluent Treated', 'effluent', 'M3')
    ]
    
    for utility_name, utility_key, unit in utilities:
        utility_data = extract_utility_balance_data(month, year, calculation_result, utility_name, utility_key)
        
        ws[f'A{row}'] = utility_name
        ws[f'A{row}'].font = BOLD_FONT
        ws[f'A{row}'].fill = SUBSECTION_FILL
        row += 1
        
        # Total Demand
        ws[f'A{row}'] = "  Total Demand"
        ws[f'B{row}'] = round(utility_data['demand']['total_demand'], 2)
        ws[f'C{row}'] = unit
        ws[f'A{row}'].font = BOLD_FONT
        row += 1
        
        # Total Supply
        ws[f'A{row}'] = "  Total Supply"
        ws[f'B{row}'] = round(utility_data['supply']['total_supply'], 2)
        ws[f'C{row}'] = unit
        ws[f'A{row}'].font = BOLD_FONT
        row += 1
        
        # Balance
        ws[f'A{row}'] = "  Imbalance"
        ws[f'B{row}'] = round(utility_data['balance'], 2)
        ws[f'C{row}'] = unit
        ws[f'A{row}'].font = BOLD_FONT
        row += 2
    
    return row


# ============================================================
# SECTION IV: OTHER UTILITIES BALANCE
# ============================================================

def get_utility_norm_from_db(
    month: int,
    year: int,
    plant_name: str,
    utility_name: str,
    material_name: str,
    account_name: str = "Utilities",
    issuing_plant_name: str = None,
) -> float:
    """
    Fetch norm value from NormsMonthDetail table for a specific plant, utility, and material.
    
    Args:
        month: Month number (1-12)
        year: Year
        plant_name: Plant name (e.g., 'NMD - Utility Plant')
        utility_name: Utility name (e.g., 'Boiler Feed Water')
        material_name: Material name (e.g., 'D M Water', 'Power_Dis')
    
    Returns:
        Norm value from database, or 0.0 if not found
    """
    try:
        return get_month_norm(
            month=month,
            year=year,
            plant_name=plant_name,
            utility_name=utility_name,
            material_name=material_name,
            account_name=account_name,
            issuing_plant_name=issuing_plant_name,
            required=False,
        )
    except Exception as e:
        print(f"  [NORM FETCH] Error fetching norm for {plant_name}/{utility_name}/{material_name}: {e}")
        return 0.0


# ============================================================
# SECTION V: CHEMICAL BALANCE
# ============================================================

def write_chemical_balance_section(ws, start_row: int, month: int, year: int, calculation_result: dict) -> int:
    """Write Section V: Chemical Balance with individual balance for each chemical."""
    row = start_row
    
    # Section header
    for col in range(1, 6):
        ws.cell(row=row, column=col).fill = SECTION_FILL
    ws.merge_cells(f'A{row}:E{row}')
    header_cell = ws[f'A{row}']
    header_cell.value = "SECTION V: CHEMICAL BALANCE"
    header_cell.font = SECTION_FONT
    header_cell.alignment = Alignment(horizontal='left', vertical='center')
    row += 1
    
    # List of all chemicals to create balances for
    # Format: (chemical_name, uom, parent_utility, parent_utility_name, consumer_name)
    # Norms will be fetched from database
    chemicals = [
        # BFW Chemicals
        ('CHEM CYCLO HEXY', 'KG', 'bfw', 'Boiler Feed Water', 'BFW Plant'),
        ('CHEM MORPHOLENE', 'MT', 'bfw', 'Boiler Feed Water', 'BFW Plant'),
        ('KEM WATREAT B 70M', 'KG', 'bfw', 'Boiler Feed Water', 'BFW Plant'),
        
        # DM Water Chemicals  (MaterialName must match DB exactly; UtilityName = 'D M Water')
        ('CAUSTIC SODA LYE – GRADE 1',              'MT', 'dm', 'D M Water', 'DM Water Plant'),
        ('CHEM ALUM.SULFATE, AL2(SO4)3,18H2O',     'KG', 'dm', 'D M Water', 'DM Water Plant'),
        ('CHEM  SODIUM SULPHITE;PN:MIS 19OX',       'KG', 'dm', 'D M Water', 'DM Water Plant'),
        ('POLYELECTROLYTE',                         'KG', 'dm', 'D M Water', 'DM Water Plant'),
        ('SODIUM CHLORIDE IS 797 GRADE1',           'MT', 'dm', 'D M Water', 'DM Water Plant'),
        
        # HRSG Chemicals (will handle separately due to multiple HRSGs)
        ('CHEM TRISODIUM PHOSPHATE', 'KG', 'hrsg', 'HRSG1_SHP STEAM', 'HRSG Plants'),
    ]
    
    # Write balance for each chemical
    for chemical_name, uom, parent_utility, parent_utility_name, consumer_name in chemicals:
        row = write_single_chemical_balance(ws, row, month, year, calculation_result, 
                                           chemical_name, uom, parent_utility, parent_utility_name, consumer_name)
        row += 1  # Add spacing between chemicals
    
    return row


def write_single_chemical_balance(ws, start_row: int, month: int, year: int, calculation_result: dict,
                                  chemical_name: str, uom: str, parent_utility: str, parent_utility_name: str, 
                                  consumer_name: str) -> int:
    """
    Write a single chemical balance with 5-column layout (Demand, Norm, Quantity, Supply Source, Quantity).
    
    Args:
        chemical_name: Name of the chemical (e.g., 'CHEM CYCLO HEXY')
        uom: Unit of measurement (e.g., 'KG', 'MT')
        parent_utility: Parent utility that consumes this chemical ('bfw', 'dm', 'hrsg')
        parent_utility_name: Utility name for database lookup (e.g., 'Boiler Feed Water')
        consumer_name: Name of the consumer (e.g., 'BFW Plant')
    """
    utility_consumption = calculation_result.get('utility_consumption', {})
    
    # Fetch norm from database
    plant_name = 'NMD - Utility Plant'
    norm = get_utility_norm_from_db(month, year, plant_name, parent_utility_name, chemical_name)
    
    row = start_row
    
    # Subsection header
    for col in range(1, 6):
        ws.cell(row=row, column=col).fill = SUBSECTION_FILL
    ws.merge_cells(f'A{row}:E{row}')
    header_cell = ws[f'A{row}']
    display_name = (chemical_name
                    .replace(' – GRADE 1', '')
                    .replace(', AL2(SO4)3,18H2O', '')
                    .replace(';PN:MIS 19OX', '')
                    .replace(' IS 797 GRADE1', '')
                    .strip())
    header_cell.value = f"{display_name} BALANCE"
    header_cell.font = Font(bold=True, size=11)
    header_cell.alignment = Alignment(horizontal='left', vertical='center')
    row += 1
    
    # Column headers
    headers = ['Demand', 'Norm', f'Quantity {uom}', 'Supply Source', f'Quantity {uom}']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = BOLD_FONT
        cell.fill = SUBSECTION_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    start_data_row = row
    
    # Calculate chemical quantity based on parent utility consumption
    chemical_quantity = 0.0
    parent_utility_quantity = 0.0
    parent_utility_uom = ''
    
    if parent_utility == 'bfw':
        # BFW chemicals - based on total BFW consumption
        # BFW data structure: {'for_hrsg_m3': X, 'for_hp_prds_m3': X, ..., 'total_m3': X}
        bfw_data = utility_consumption.get('bfw', {})
        parent_utility_quantity = bfw_data.get('total_m3', 0.0)
        parent_utility_uom = 'M3'
        chemical_quantity = parent_utility_quantity * norm
        
    elif parent_utility == 'dm':
        # DM Water chemicals - based on total DM consumption
        # DM data structure: {'for_bfw_m3': X, 'process_m3': X, ..., 'total_m3': X}
        dm_data = utility_consumption.get('dm_water', {})
        parent_utility_quantity = dm_data.get('total_m3', 0.0)
        parent_utility_uom = 'M3'
        chemical_quantity = parent_utility_quantity * norm
        
    elif parent_utility == 'hrsg':
        # HRSG chemicals - based on total HRSG SHP consumption
        # Sum all HRSG SHP generation (free + supplementary)
        hrsg1_shp = utility_consumption.get('shp_from_hrsg1', 0.0)
        hrsg2_shp = utility_consumption.get('shp_from_hrsg2', 0.0)
        hrsg3_shp = utility_consumption.get('shp_from_hrsg3', 0.0)
        parent_utility_quantity = hrsg1_shp + hrsg2_shp + hrsg3_shp
        parent_utility_uom = 'MT'
        chemical_quantity = parent_utility_quantity * norm
    
    # LEFT SIDE: DEMAND
    # Utility for Utility (U4U)
    ws[f'A{row}'] = "Utility for Utility (U4U)"
    ws[f'A{row}'].font = BOLD_FONT
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = THIN_BORDER
    row += 1
    
    ws[f'A{row}'] = f"  {consumer_name}"
    ws[f'B{row}'] = norm
    ws[f'C{row}'] = round(chemical_quantity, 2)
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = THIN_BORDER
    
    # Total Demand
    ws[f'A{row}'] = "Total Demand"
    ws[f'C{row}'] = round(chemical_quantity, 2)
    ws[f'A{row}'].font = BOLD_FONT
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = THIN_BORDER
        ws.cell(row=row, column=col).fill = TOTAL_FILL
    row += 1
    
    # Only show supply and balance if there's actual demand/supply (chemical_quantity > 0)
    if chemical_quantity > 0:
        # RIGHT SIDE: SUPPLY
        supply_row = start_data_row
        
        # If HRSG chemical, show breakdown by HRSG
        if parent_utility == 'hrsg':
            ws[f'D{supply_row}'] = "Supply by HRSG"
            ws[f'D{supply_row}'].font = BOLD_FONT
            for col in range(1, 6):
                ws.cell(row=supply_row, column=col).border = THIN_BORDER
            supply_row += 1
            
            # HRSG1
            hrsg1_shp = utility_consumption.get('shp_from_hrsg1', 0.0)
            hrsg1_chemical = hrsg1_shp * norm
            ws[f'D{supply_row}'] = "  HRSG-1"
            ws[f'E{supply_row}'] = round(hrsg1_chemical, 2)
            for col in range(1, 6):
                ws.cell(row=supply_row, column=col).border = THIN_BORDER
            supply_row += 1
            
            # HRSG2
            hrsg2_shp = utility_consumption.get('shp_from_hrsg2', 0.0)
            hrsg2_chemical = hrsg2_shp * norm
            ws[f'D{supply_row}'] = "  HRSG-2"
            ws[f'E{supply_row}'] = round(hrsg2_chemical, 2)
            for col in range(1, 6):
                ws.cell(row=supply_row, column=col).border = THIN_BORDER
            supply_row += 1
            
            # HRSG3
            hrsg3_shp = utility_consumption.get('shp_from_hrsg3', 0.0)
            hrsg3_chemical = hrsg3_shp * norm
            ws[f'D{supply_row}'] = "  HRSG-3"
            ws[f'E{supply_row}'] = round(hrsg3_chemical, 2)
            for col in range(1, 6):
                ws.cell(row=supply_row, column=col).border = THIN_BORDER
            supply_row += 1
        else:
            # For BFW and DM chemicals, show single supply source
            ws[f'D{supply_row}'] = "Chemical Supply"
            ws[f'D{supply_row}'].font = BOLD_FONT
            for col in range(1, 6):
                ws.cell(row=supply_row, column=col).border = THIN_BORDER
            supply_row += 1
            
            ws[f'D{supply_row}'] = "  Purchased/Supplied"
            ws[f'E{supply_row}'] = round(chemical_quantity, 2)
            for col in range(1, 6):
                ws.cell(row=supply_row, column=col).border = THIN_BORDER
            supply_row += 1
        
        # Total Supply (align with Total Demand row)
        ws[f'D{row}'] = "Total Supply"
        ws[f'E{row}'] = round(chemical_quantity, 2)
        ws[f'D{row}'].font = BOLD_FONT
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).fill = TOTAL_FILL
        
        row += 1
        
        # Balance check
        balance = 0.0  # Should always be 0 for chemicals (demand = supply)
        ws[f'A{row}'] = "Balance (Supply - Demand)"
        ws[f'C{row}'] = round(balance, 2)
        ws[f'A{row}'].font = Font(bold=True, color="FF0000" if abs(balance) > 0.01 else "000000")
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1
    else:
        # No supply case - just move to next row after Total Demand
        row += 1
    
    return row
