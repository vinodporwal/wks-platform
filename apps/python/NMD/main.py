"""
NMD CPP — Step 1: Fetch & Display Asset Details for a Month

Interactive entry point. Asks month + year, then fetches and displays:
  - Plant info
  - Power generation assets (PowerGenerationAssets)
  - Asset snapshot (OperationalHours + AssetAvailability for the month)
  - Steam generation assets (SteamGenerationAssets)
  - Demand segregation (Process + Fixed)
  - Import power

No calculation, no dispatch, no U4U loop — just data fetching and display.

Usage:
    python main.py
    python main.py --month 4 --year 2026
"""

from __future__ import annotations

import argparse
import sys
import os
from datetime import datetime

# Ensure the NMD module root is on sys.path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from plant_mapper import NMD_PLANT_ID, PLANT_REGISTRY
from engine.report_logger import setup_logging, LogCapture, save_text_log
from database.queries import (
    fetch_power_generation_assets,
    fetch_asset_availability_with_hours,
    fetch_steam_generation_assets,
    fetch_steam_asset_operational_hours,
    fetch_consolidated_demand_with_uom,
    fetch_import_power,
    calculate_free_steam_from_gt,
    fetch_generation_utilities_and_u4u,
    fetch_cpp_norms_for_utilities,
)
from engine.u4u_iteration_loop import run_u4u_iteration
from engine.norms_reader import NMDNormsReader
from engine.bpc_ods_reader import BPCODSReader
from services.norms_save_service import save_calculated_norms
from services.cost_module_service import (
    calculate_and_print_utility_prices,
    build_bpc_comparison_table_text,
    build_yearly_bpc_comparison_table_text,
)

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
LOG_FOLDER = os.path.join(_SCRIPT_DIR, "logs")

_MONTH_NAMES = {
    1: "January", 2: "February", 3: "March", 4: "April",
    5: "May", 6: "June", 7: "July", 8: "August",
    9: "September", 10: "October", 11: "November", 12: "December",
}


def _sec(title: str):
    print("\n" + "=" * 70)
    print(f"  {title}")
    print("=" * 70)


# ---------------------------------------------------------------------------
# Display functions
# ---------------------------------------------------------------------------

def _pct_diff(actual: float, target: float) -> float:
    """Calculate percentage difference: (actual - target) / target * 100"""
    if target == 0:
        return 0.0
    return (actual - target) / target * 100


def _generate_u4u_comparison_excel(
    detail_records: list,
    bpc_reader,
    log_file_path: str,
):
    """Generate Excel file for U4U CONSUMPTION vs BPC COMPARISON TABLE."""
    if not OPENPYXL_AVAILABLE:
        print("  openpyxl not available - skipping Excel generation")
        return

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "U4U Comparison"

        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')

        # Headers
        headers = [
            "Plant Name", "Generation Utility", "UOM", "Gen Qty", "BPC Gen Qty", "Gen Diff %",
            "U4U Plant", "U4U Utility", "UOM", "Norm", "Consumption Qty", "BPC Qty", "Qty Diff %", "Credit"
        ]
        ws.append(headers)

        # Style header row
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_align

        # Data rows
        for rec in sorted(detail_records,
                          key=lambda r: (r.get("producer", ""), r.get("producer_utility", ""),
                                         r.get("material", ""))):
            producer = rec.get("producer", "")
            producer_utility = rec.get("producer_utility", "")
            material = rec.get("material", "")
            norm = rec.get("norm", 0.0)
            quantity = rec.get("quantity", 0.0)
            generation = rec.get("generation", 0.0)
            is_credit = rec.get("is_credit", False)

            # BPC keys
            if producer_utility == "POWERGEN":
                bpc_gen_key = producer
                bpc_qty_key = producer
            elif material == "POWERGEN":
                bpc_gen_key = producer_utility
                bpc_qty_key = rec.get("issuing_plant", "") or producer_utility
            else:
                bpc_gen_key = producer_utility
                bpc_qty_key = producer_utility

            bpc_gen = bpc_reader.get_bpc_gen_qty_for_producer(bpc_gen_key)
            bpc_qty = bpc_reader.get_bpc_qty_for_producer_material(bpc_qty_key, material)

            gen_diff = _pct_diff(generation, bpc_gen)
            qty_diff = _pct_diff(quantity, bpc_qty)

            producer_uom = rec.get("material_uom", "")
            u4u_plant = rec.get("issuing_plant", "") or rec.get("producer", "")
            credit_flag = "Yes" if is_credit else "No"

            row = [
                producer, producer_utility, "", generation, bpc_gen, gen_diff,
                u4u_plant, material, producer_uom, norm, quantity, bpc_qty, qty_diff, credit_flag
            ]
            ws.append(row)

            # Style data row
            for col_num in range(1, len(row) + 1):
                cell = ws.cell(row=ws.max_row, column=col_num)
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')

        # Auto-fit columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = "A2"

        # Save Excel file
        excel_path = log_file_path.replace(".log", "_U4U_Comparison.xlsx")
        wb.save(excel_path)
        print(f"  Excel comparison table saved: {excel_path}")

    except Exception as e:
        print(f"  Error generating Excel file: {e}")


def _print_plant_info():
    _sec("PLANT INFO")
    plant = PLANT_REGISTRY.get(NMD_PLANT_ID, {})
    print(f"  plant_id:     {NMD_PLANT_ID}")
    print(f"  name:         {plant.get('name', '?')}")
    print(f"  display_name: {plant.get('display_name', '?')}")
    print(f"  short_code:   {plant.get('short_code', '?')}")


def _print_power_assets(plant_id: str):
    _sec("POWER GENERATION ASSETS  (PowerGenerationAssets)")
    rows = fetch_power_generation_assets(plant_id)
    if not rows:
        print("  No assets found")
        return
    print(f"  {len(rows)} asset(s):")
    print(f"  {'Asset':<30}  {'Type':<8}  {'PlantCode':<12}")
    print(f"  {'-'*30}  {'-'*8}  {'-'*12}")
    for a in rows:
        print(f"  {a['asset_name']:<30}  {a['asset_type']:<8}  {a.get('plant_code', '-'):<12}")


def _print_asset_snapshot(plant_id: str, month: int, year: int):
    _sec(f"ASSET SNAPSHOT  ({_MONTH_NAMES.get(month, month)} {year})")
    rows = fetch_asset_availability_with_hours(month, year, plant_id)
    if not rows:
        print("  No assets found for this period")
        return
    print(f"  {len(rows)} asset(s)")
    print(f"  {'Asset':<28}  {'Type':<6}  {'Hours':>7}  {'MinMW':>7}  {'MaxMW':>7}  {'FixMin':>7}  {'FixMax':>7}  {'Prio':>5}  {'Avail':>6}")
    print(f"  {'-'*28}  {'-'*6}  {'-'*7}  {'-'*7}  {'-'*7}  {'-'*7}  {'-'*7}  {'-'*5}  {'-'*6}")
    for r in rows:
        hrs = r.get("operational_hours", 0)
        min_mw = r.get("min_operating_capacity") or 0
        max_mw = r.get("max_operating_capacity") or 0
        fix_min = r.get("fixed_min") or 0
        fix_max = r.get("fixed_max") or 0
        pri = r.get("priority")
        pri_str = str(pri) if pri is not None else "N/A"
        avail = "Yes" if r.get("is_available") else "No"
        print(f"  {r['asset_name']:<28}  {r.get('asset_type', '-'):<6}  {hrs:>7.1f}  {min_mw:>7.2f}  {max_mw:>7.2f}  {fix_min:>7.2f}  {fix_max:>7.2f}  {pri_str:>5}  {avail:>6}")


def _print_gt_heat_rate_and_free_steam(plant_id: str, month: int, year: int):
    _sec(f"GT HEAT RATE & FREE STEAM  ({_MONTH_NAMES.get(month, month)} {year})")
    power_assets = fetch_asset_availability_with_hours(month, year, plant_id)
    if not power_assets:
        print("  No power assets found for this period")
        return

    result = calculate_free_steam_from_gt(power_assets, month, year)
    details = result["details"]
    total_fs = result["total_free_steam_mt"]

    print(f"  {'GT Asset':<24}  {'Curve':<6}  {'LoadMW':>7}  {'Hours':>7}  {'GrossMWh':>10}  {'HeatRate':>9}  {'FSFactor':>9}  {'FreeSteamMT':>12}  {'→ HRSG':<10}")
    print(f"  {'-'*24}  {'-'*6}  {'-'*7}  {'-'*7}  {'-'*10}  {'-'*9}  {'-'*9}  {'-'*12}  {'-'*10}")
    for asset_name, d in details.items():
        print(f"  {asset_name:<24}  {d['gt_name']:<6}  {d['load_mw']:>7.2f}  {d['operational_hours']:>7.1f}  {d['gross_mwh']:>10.2f}  {d['heat_rate_kcal_kwh']:>9.2f}  {d['free_steam_factor']:>9.4f}  {d['free_steam_mt']:>12.2f}  {d['linked_hrsg']:<10}")
    print(f"  {'-'*24}  {'-'*6}  {'-'*7}  {'-'*7}  {'-'*10}  {'-'*9}  {'-'*9}  {'-'*12}  {'-'*10}")
    print(f"  {'TOTAL FREE STEAM':<24}  {'':6}  {'':7}  {'':7}  {'':10}  {'':9}  {'':9}  {total_fs:>12.2f}")
    print(f"\n  Free Steam (MT) = GrossMWh × FreeSteamFactor (from CPP_GTHeatRate at GT load)")
    print(f"  This free steam is transferred to the respective linked HRSG as SHP steam input.")


def _print_generation_utilities_and_u4u(month: int, year: int):
    _sec(f"GENERATION UTILITIES & U4U CONSUMPTION MATRIX  ({_MONTH_NAMES.get(month, month)} {year})")
    data = fetch_generation_utilities_and_u4u(month, year)
    gen_utils = data["generation_utilities"]
    if not gen_utils:
        print("  No norms data found for NMD sub-plants for this period")
        return

    producers = data["producer_utilities"]
    consumers = data["consumer_utilities"]
    all_utils = data["all_utilities"]

    print(f"  Producer Utilities ({len(producers)}): {', '.join(producers)}")
    print(f"  Consumer/U4U Utilities ({len(consumers)}): {', '.join(consumers)}")
    print(f"  Total Unique Utilities: {len(all_utils)}")

    print(f"\n  {'Plant':<28}  {'Utility (Producer)':<22}  {'UOM':<6}  {'Generation':>14}  {'U4U Consumes':<50}")
    print(f"  {'-'*28}  {'-'*22}  {'-'*6}  {'-'*14}  {'-'*50}")
    for g in gen_utils:
        consumes_str = "; ".join(
            f"{c['material']}({c['norm']:.4f})" for c in g["consumes"]
        )
        if len(consumes_str) > 50:
            consumes_str = consumes_str[:47] + "..."
        print(f"  {g['plant_name']:<28}  {g['utility_name']:<22}  {g['uom']:<6}  {g['generation']:>14.2f}  {consumes_str:<50}")

    print(f"\n  --- U4U CONSUMPTION DETAILS ---\n")
    print(f"  {'Producer Utility':<22}  {'Consumes (Material)':<22}  {'Norm':>10}  {'Qty':>14}  {'UOM':<6}  {'Issuing Plant':<28}")
    print(f"  {'-'*22}  {'-'*22}  {'-'*10}  {'-'*14}  {'-'*6}  {'-'*28}")
    u4u = data["u4u_matrix"]
    for prod_util in sorted(u4u.keys()):
        for material in sorted(u4u[prod_util].keys()):
            entry = u4u[prod_util][material]
            print(f"  {prod_util:<22}  {material:<22}  {entry['norm']:>10.6f}  {entry['qty']:>14.2f}  {entry['uom']:<6}  {entry['issuing_plant']:<28}")

    print(f"\n  Note: Norm = consumption per unit generation. Qty = actual consumption for the month.")
    print(f"  U4U consumption adds to total demand for each consumed utility.")


def _print_cpp_norms_for_u4u(month: int, year: int):
    _sec(f"CPPNORMS — U4U CONSUMPTION NORMS  ({_MONTH_NAMES.get(month, month)} {year})")
    rows = fetch_cpp_norms_for_utilities(month, year)
    if not rows:
        print("  No CPPNorms data found for NMD sub-plants for this period")
        return

    if month >= 4:
        fy = f"{year}-{str(year + 1)[-2:]}"
    else:
        fy = f"{year - 1}-{str(year)[-2:]}"

    month_col = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}.get(month, "Apr")
    print(f"  Financial Year: {fy}  |  Norm Column: {month_col}_Norms  |  Total Rows: {len(rows)}")

    # Group by (plant_name, utility_name) → list of U4U norms
    grouped = {}
    for r in rows:
        key = (r["plant_name"], r["utility_name"])
        if key not in grouped:
            grouped[key] = {
                "utility_uom": r["utility_uom"],
                "u4u_list": [],
            }
        if r["material_name"] and r["material_name"] != "No Material":
            grouped[key]["u4u_list"].append({
                "material": r["material_name"],
                "norm": r["norm"],
                "uom": r["issuing_uom"],
                "issuing_plant": r["issuing_plant"],
            })

    print(f"\n  {'Plant':<28}  {'Generation Utility':<22}  {'UOM':<6}  {'U4U Material':<22}  {'Norm':>12}  {'UOM':<6}  {'Issuing Plant':<28}")
    print(f"  {'-'*28}  {'-'*22}  {'-'*6}  {'-'*22}  {'-'*12}  {'-'*6}  {'-'*28}")

    for (plant, utility), info in sorted(grouped.items()):
        u4u_list = info["u4u_list"]
        if not u4u_list:
            print(f"  {plant:<28}  {utility:<22}  {info['utility_uom']:<6}  {'(no U4U consumes)':<22}  {'':>12}  {'':<6}  {'':<28}")
            continue
        first = True
        for u in u4u_list:
            if first:
                print(f"  {plant:<28}  {utility:<22}  {info['utility_uom']:<6}  {u['material']:<22}  {u['norm']:>12.6f}  {u['uom']:<6}  {u['issuing_plant']:<28}")
                first = False
            else:
                print(f"  {'':28}  {'':22}  {'':6}  {u['material']:<22}  {u['norm']:>12.6f}  {u['uom']:<6}  {u['issuing_plant']:<28}")

    print(f"\n  Norm = U4U consumption per unit of generation utility.")
    print(f"  Consumption QTY = Norm × Generation QTY (from NormsMonthDetail).")


def _print_u4u_bpc_comparison(u4u_result: dict, month: int, year: int, log_file_path: str = None):
    """
    Print a comparison table showing model-calculated U4U consumption
    vs BPC (from BPC.ods) values.

    Columns:
      Plant Name, Generation Utility, UOM, Gen Qty, BPC Gen Qty, Gen Diff %,
      U4U Plant, U4U Utility, UOM, Norm, Consumption Qty, BPC Qty, Qty Diff %
    """
    import logging
    log = logging.getLogger(__name__)

    _sec(f"U4U vs BPC COMPARISON  ({_MONTH_NAMES.get(month, month)} {year})")

    bpc_reader = BPCODSReader.get_reader(month, year)
    bpc_reader.load()
    if not bpc_reader.is_available:
        print("  BPC.ods not available — skipping comparison")
        return

    bpc_reader.log_bpc_summary()

    bpc_gen_map = bpc_reader.get_bpc_generation_quantities()
    bpc_qty_map = bpc_reader.get_bpc_quantities()

    detail_records = u4u_result.get("detail_records", [])
    if not detail_records:
        print("  No U4U detail records from iteration loop — skipping comparison")
        return

    def _pct_diff(model_val: float, bpc_val: float) -> float:
        if abs(bpc_val) < 1e-10:
            return 0.0 if abs(model_val) < 1e-10 else 100.0
        return (model_val - bpc_val) / abs(bpc_val) * 100.0

    # Power balance summary
    pr = u4u_result.get("power_result", {}) or {}
    gt_stg_gen = float(pr.get("total_generation_mwh", 0.0))
    import_pw = pr.get("import_power", {})
    import_mwh = float(import_pw.get("total_mwh", 0.0)) if isinstance(import_pw, dict) and import_pw.get("success") else 0.0
    total_supply = gt_stg_gen + import_mwh
    total_demand = float(pr.get("demand_mwh", 0.0))
    print("")
    print("  POWER BALANCE SUMMARY")
    print(f"  {'Total Power Demand':<30}  {total_demand:>14.2f} MWh")
    print(f"  {'GT/STG Generation':<30}  {gt_stg_gen:>14.2f} MWh")
    print(f"  {'Import Power':<30}  {import_mwh:>14.2f} MWh")
    print(f"  {'Total Supply (Gen + Import)':<30}  {total_supply:>14.2f} MWh")
    print(f"  {'Surplus/Deficit':<30}  {total_supply - total_demand:>14.2f} MWh")

    # Build the comparison table
    print("")
    print("  " + "=" * 180)
    print("  U4U CONSUMPTION vs BPC COMPARISON TABLE")
    print("  " + "=" * 180)
    print(f"  {'Plant Name':<25}  {'Generation Utility':<25}  {'UOM':<6}  {'Gen Qty':>14}  {'BPC Gen Qty':>14}  {'Gen Diff %':>10}  {'U4U Plant':<25}  {'U4U Utility':<25}  {'UOM':<6}  {'Norm':>12}  {'Consumption Qty':>14}  {'BPC Qty':>14}  {'Qty Diff %':>10}")
    print("  " + "-" * 25 + "  " + "-" * 25 + "  " + "-" * 6 + "  " + "-" * 14 + "  " + "-" * 14 + "  " + "-" * 10 + "  " + "-" * 25 + "  " + "-" * 25 + "  " + "-" * 6 + "  " + "-" * 12 + "  " + "-" * 14 + "  " + "-" * 14 + "  " + "-" * 10)

    for rec in sorted(detail_records,
                      key=lambda r: (r.get("producer", ""), r.get("producer_utility", ""),
                                     r.get("material", ""))):
        producer = rec.get("producer", "")
        producer_utility = rec.get("producer_utility", "")
        material = rec.get("material", "")
        norm = rec.get("norm", 0.0)
        quantity = rec.get("quantity", 0.0)
        generation = rec.get("generation", 0.0)
        is_credit = rec.get("is_credit", False)

        # BPC keys by plant name for POWERGEN producers, by utility name for others
        # When material is POWERGEN (Power_Dis consuming POWERGEN), key by issuing_plant
        if producer_utility == "POWERGEN":
            bpc_gen_key = producer
            bpc_qty_key = producer
        elif material == "POWERGEN":
            bpc_gen_key = producer_utility
            bpc_qty_key = rec.get("issuing_plant", "") or producer_utility
        else:
            bpc_gen_key = producer_utility
            bpc_qty_key = producer_utility

        # Look up BPC values
        bpc_gen = bpc_reader.get_bpc_gen_qty_for_producer(bpc_gen_key)
        bpc_qty = bpc_reader.get_bpc_qty_for_producer_material(bpc_qty_key, material)

        gen_diff = _pct_diff(generation, bpc_gen)
        qty_diff = _pct_diff(quantity, bpc_qty)

        # Get UOMs from the record
        producer_uom = rec.get("material_uom", "")  # material UOM
        gen_uom = ""  # We don't have producer UOM in detail records

        # For the U4U plant, use the issuing_plant from the record
        u4u_plant = rec.get("issuing_plant", "") or rec.get("producer", "")

        credit_flag = " (C)" if is_credit else ""

        print(f"  {producer[:25]:<25}  {producer_utility[:25]:<25}  {gen_uom[:6]:<6}  {generation:>14.2f}  {bpc_gen:>14.2f}  {gen_diff:>9.2f}%  {u4u_plant[:25]:<25}  {material[:25]:<25}  {producer_uom[:6]:<6}  {norm:>12.6f}  {quantity:>14.2f}  {bpc_qty:>14.2f}  {qty_diff:>9.2f}%{credit_flag}")

    print("")

    # Summary: per-producer generation comparison
    print("  GENERATION UTILITY COMPARISON SUMMARY")
    print(f"  {'Producer':<30}  {'UOM':<6}  {'Gen Qty':>14}  {'BPC Gen Qty':>14}  {'Gen Diff %':>10}  {'# Mat':>8}")
    print("  " + "-" * 30 + "  " + "-" * 6 + "  " + "-" * 14 + "  " + "-" * 14 + "  " + "-" * 10 + "  " + "-" * 8)

    producer_summary: dict = {}
    for rec in detail_records:
        p = rec.get("producer", "")
        pu = rec.get("producer_utility", "")
        bpc_key = p if pu == "POWERGEN" else pu
        if p not in producer_summary:
            bpc_gen = bpc_reader.get_bpc_gen_qty_for_producer(bpc_key)
            producer_summary[p] = {
                "utility": pu,
                "generation": rec["generation"],
                "bpc_gen": bpc_gen,
                "material_count": 0,
            }
        producer_summary[p]["material_count"] += 1

    for p in sorted(producer_summary.keys()):
        s = producer_summary[p]
        gen_diff = _pct_diff(s["generation"], s["bpc_gen"])
        print(f"  {p[:30]:<30}  {'':<6}  {s['generation']:>14.2f}  {s['bpc_gen']:>14.2f}  {gen_diff:>9.2f}%  {s['material_count']:>8d}")

    print("  " + "=" * 180)

    # Generate Excel file for comparison table
    if log_file_path:
        _generate_u4u_comparison_excel(detail_records, bpc_reader, log_file_path)


def _print_steam_assets(plant_id: str, month: int, year: int):
    _sec(f"STEAM GENERATION ASSETS  ({_MONTH_NAMES.get(month, month)} {year})")
    rows = fetch_steam_generation_assets(plant_id)
    if not rows:
        print("  No steam assets found")
        return
    ops_hrs = fetch_steam_asset_operational_hours(plant_id, month, year)
    print(f"  {len(rows)} asset(s):")
    print(f"  {'Asset':<22}  {'Type':<6}  {'Steam':<6}  {'MinMT':>7}  {'MaxMT':>7}  {'Eff':>5}  {'Prio':>5}  {'Hours':>7}  {'AlwaysOn':>9}  {'LinkedGT':<16}")
    print(f"  {'-'*22}  {'-'*6}  {'-'*6}  {'-'*7}  {'-'*7}  {'-'*5}  {'-'*5}  {'-'*7}  {'-'*9}  {'-'*16}")
    for a in rows:
        hrs = ops_hrs.get(a["asset_name"], 0.0)
        always = "Yes" if a.get("is_always_available") else "No"
        linked = a.get("linked_gt_name") or "-"
        print(f"  {a['asset_name']:<22}  {a.get('asset_type', '-'):<6}  {a.get('steam_type', '-'):<6}  {a.get('min_capacity_mt', 0):>7.1f}  {a.get('max_capacity_mt', 0):>7.1f}  {a.get('efficiency', 1.0):>5.2f}  {a.get('priority', 999):>5}  {hrs:>7.1f}  {always:>9}  {linked:<16}")


def _print_consolidated_demand(plant_id: str, month: int, year: int):
    _sec(f"CONSOLIDATED DEMAND  (Process + Fixed)  ({month}/{year})")
    rows = fetch_consolidated_demand_with_uom(plant_id, month, year)
    if not rows:
        print("  No demand data found for this period")
        return

    print(f"  {len(rows)} utility(ies)")
    print(f"  {'Utility':<28}  {'UOM':<8}  {'Process':>16}  {'Fixed':>16}  {'Total':>16}")
    print(f"  {'-'*28}  {'-'*8}  {'-'*16}  {'-'*16}  {'-'*16}")
    for r in rows:
        util = r["utility_name"]
        uom = r["uom"] or "-"
        pval = r["process_value"]
        fval = r["fixed_value"]
        total = r["total_value"]
        print(f"  {util:<28}  {uom:<8}  {pval:>16.2f}  {fval:>16.2f}  {total:>16.2f}")

    # Grand total row (informational — only for same-UOM utilities)
    print(f"  {'-'*28}  {'-'*8}  {'-'*16}  {'-'*16}  {'-'*16}")
    total_p = sum(r["process_value"] for r in rows)
    total_f = sum(r["fixed_value"] for r in rows)
    total_t = sum(r["total_value"] for r in rows)
    print(f"  {'GRAND TOTAL':<28}  {'':8}  {total_p:>16.2f}  {total_f:>16.2f}  {total_t:>16.2f}")
    print(f"\n  Note: Grand total mixes UOMs — for reference only.")


def _print_import_power(plant_id: str, month: int, year: int):
    _sec(f"IMPORT POWER  ({month}/{year})")
    imp = fetch_import_power(plant_id, month, year)
    total = imp.get("total_mwh", 0.0)
    sources = imp.get("sources", [])
    print(f"  Total Import Power: {total:.2f} MWh  ({len(sources)} sources)")
    if sources:
        print(f"  {'Source':<25}  {'MWh':>12}  {'Capacity':>10}  {'Hours':>8}")
        print(f"  {'-'*25}  {'-'*12}  {'-'*10}  {'-'*8}")
        for s in sources:
            print(f"  {s.get('source_name', '?'):<25}  {s.get('mwh', 0):>12.2f}  {s.get('capacity_mw', 0):>10.2f}  {s.get('hours', 0):>8.1f}")


def _print_supply_vs_demand(plant_id: str, month: int, year: int):
    _sec(f"SUPPLY vs DEMAND ANALYSIS  ({_MONTH_NAMES.get(month, month)} {year})")

    # ── Fetch all data ───────────────────────────────────────────────────
    demand_rows = fetch_consolidated_demand_with_uom(plant_id, month, year)
    power_assets = fetch_asset_availability_with_hours(month, year, plant_id)
    steam_assets = fetch_steam_generation_assets(plant_id)
    steam_hrs = fetch_steam_asset_operational_hours(plant_id, month, year)

    # ── Build demand lookup by utility name ───────────────────────────────
    demand_by_name = {r["utility_name"]: r for r in demand_rows}

    # ====================================================================
    # POWER BALANCE
    # ====================================================================
    print("\n  --- POWER BALANCE ---\n")

    # Demand side
    power_process_kwh = 0.0
    power_fixed_kwh = 0.0
    for r in demand_rows:
        name = r["utility_name"].upper().strip()
        if name in ("POWER_DIS", "POWER"):
            power_process_kwh += r["process_value"]
            power_fixed_kwh += r["fixed_value"]

    total_demand_mwh = (power_process_kwh + power_fixed_kwh) / 1000.0

    print(f"  {'DEMAND':<30}  {'MWh':>14}")
    print(f"  {'-'*30}  {'-'*14}")
    print(f"  {'Power (Process)':<30}  {power_process_kwh / 1000.0:>14.2f}")
    print(f"  {'Power (Fixed)':<30}  {power_fixed_kwh / 1000.0:>14.2f}")
    print(f"  {'Total Power Demand':<30}  {total_demand_mwh:>14.2f}")

    # Import power (subtracted from demand)
    imp = fetch_import_power(plant_id, month, year)
    import_mwh = imp.get("total_mwh", 0.0)
    import_sources = imp.get("per_source", [])

    if import_mwh > 0 or import_sources:
        print(f"\n  {'IMPORT POWER (reduces demand)':<30}  {'Cap MW':>8}  {'Hours':>8}  {'MWh':>14}")
        print(f"  {'-'*30}  {'-'*8}  {'-'*8}  {'-'*14}")
        for s in import_sources:
            print(f"  {s['source_name']:<30}  {s['capacity_mw']:>8.2f}  {s['hours']:>8.1f}  {s['mwh']:>14.2f}")
        print(f"  {'-'*30}  {'-'*8}  {'-'*8}  {'-'*14}")
        print(f"  {'Total Import Power':<30}  {'':>8}  {'':>8}  {import_mwh:>14.2f}")

    net_demand_mwh = total_demand_mwh - import_mwh
    print(f"\n  {'Net Demand (after import)':<30}  {net_demand_mwh:>14.2f}")

    # Supply side (GT + STG only)
    print(f"\n  {'SUPPLY (GT + STG Assets)':<30}  {'Max MW':>8}  {'Hours':>8}  {'MWh':>14}")
    print(f"  {'-'*30}  {'-'*8}  {'-'*8}  {'-'*14}")

    total_supply_mwh = 0.0
    for a in power_assets:
        hrs = a.get("operational_hours", 0) or 0
        max_mw = a.get("max_operating_capacity") or 0
        is_avail = a.get("is_available", False)
        asset_mwh = max_mw * hrs if is_avail else 0.0
        total_supply_mwh += asset_mwh
        status = "OK" if is_avail else "DOWN"
        print(f"  {a['asset_name']:<30}  {max_mw:>8.2f}  {hrs:>8.1f}  {asset_mwh:>14.2f}  [{status}]")

    print(f"  {'-'*30}  {'-'*8}  {'-'*8}  {'-'*14}")
    print(f"  {'Total Power Supply':<30}  {'':>8}  {'':>8}  {total_supply_mwh:>14.2f}")

    # Balance
    power_balance = total_supply_mwh - net_demand_mwh
    pct = (total_supply_mwh / net_demand_mwh * 100.0) if net_demand_mwh > 0 else 0.0
    status = "SURPLUS" if power_balance >= 0 else "DEFICIT"
    print(f"\n  {'Power Balance (Supply - Net Demand)':<30}  {power_balance:>14.2f} MWh  [{status}]")
    print(f"  {'Supply / Net Demand':<30}  {pct:>14.1f} %")
    if power_balance < 0:
        print(f"  ** WARNING: Power supply CANNOT meet net demand by {abs(power_balance):.2f} MWh **")
    else:
        print(f"  ** OK: Power supply can meet net demand with {power_balance:.2f} MWh surplus **")

    # ====================================================================
    # STEAM BALANCE (SHP, HP, MP, LP)
    # ====================================================================
    print("\n\n  --- STEAM BALANCE ---\n")

    # Map steam demand from consolidated demand rows
    _STEAM_DEMAND_KEYS = {
        "SHP": ["SHP Steam_Dis", "SHP Steam", "SHP_STEAM"],
        "HP":  ["HP Steam_Dis", "HP Steam", "HP_STEAM"],
        "MP":  ["MP Steam_Dis", "MP Steam", "MP_STEAM"],
        "LP":  ["LP Steam_Dis", "LP Steam", "LP_STEAM"],
    }

    for steam_type in ("SHP", "HP", "MP", "LP"):
        print(f"\n  === {steam_type} STEAM ===\n")

        # Demand
        steam_demand_mt = 0.0
        for key in _STEAM_DEMAND_KEYS[steam_type]:
            r = demand_by_name.get(key)
            if r:
                steam_demand_mt += r["process_value"] + r["fixed_value"]

        print(f"  {'DEMAND':<30}  {'MT':>14}")
        print(f"  {'-'*30}  {'-'*14}")
        print(f"  {'Total {} Steam Demand'.format(steam_type):<30}  {steam_demand_mt:>14.2f}")

        # Supply
        print(f"\n  {'SUPPLY (Available Assets)':<30}  {'MaxMT/hr':>10}  {'Hours':>8}  {'MT':>14}")
        print(f"  {'-'*30}  {'-'*10}  {'-'*8}  {'-'*14}")

        total_steam_supply = 0.0
        for a in steam_assets:
            a_steam = (a.get("steam_type") or "").upper()
            a_type = (a.get("asset_type") or "").upper()

            # Match steam type: SHP matches SHP, PRDS matches its steam type, STG matches ALL
            matches = False
            if steam_type == "SHP" and a_type == "HRSG":
                matches = True
            elif a_type == "PRDS" and a_steam == steam_type:
                matches = True
            elif a_type == "STG" and steam_type == "SHP":
                # STG consumes SHP, doesn't produce it — skip for supply
                matches = False

            if not matches:
                continue

            hrs = steam_hrs.get(a["asset_name"], 0.0)
            max_mt = a.get("max_capacity_mt", 0) or 0
            always_on = a.get("is_always_available", False)
            asset_mt = max_mt * hrs

            # Cap PRDS at demand (it's "unlimited" at 999999)
            if a_type == "PRDS" and max_mt >= 999999:
                asset_mt = steam_demand_mt  # PRDS can meet whatever is needed

            total_steam_supply += asset_mt
            status = "ALWAYS" if always_on else ("OK" if hrs > 0 else "DOWN")
            print(f"  {a['asset_name']:<30}  {max_mt:>10.1f}  {hrs:>8.1f}  {asset_mt:>14.2f}  [{status}]")

        print(f"  {'-'*30}  {'-'*10}  {'-'*8}  {'-'*14}")
        print(f"  {'Total {} Steam Supply'.format(steam_type):<30}  {'':>10}  {'':>8}  {total_steam_supply:>14.2f}")

        # Balance
        steam_balance = total_steam_supply - steam_demand_mt
        pct_s = (total_steam_supply / steam_demand_mt * 100.0) if steam_demand_mt > 0 else 0.0
        status_s = "SURPLUS" if steam_balance >= 0 else "DEFICIT"
        print(f"\n  {'{} Steam Balance'.format(steam_type):<30}  {steam_balance:>14.2f} MT  [{status_s}]")
        print(f"  {'Supply / Demand':<30}  {pct_s:>14.1f} %")
        if steam_balance < 0:
            print(f"  ** WARNING: {steam_type} steam supply CANNOT meet demand by {abs(steam_balance):.2f} MT **")
        else:
            print(f"  ** OK: {steam_type} steam supply can meet demand with {steam_balance:.2f} MT surplus **")


# ---------------------------------------------------------------------------
# Single Month Execution
# ---------------------------------------------------------------------------

def execute_single_month(month: int, year: int, plant_id: str, plant_name: str, log_folder: str,
                         verbose: bool = True, dry_run: bool = False):
    """
    Execute NMD model for a single month.
    
    Args:
        month: Month (1-12)
        year: Calendar year
        plant_id: Plant ID
        plant_name: Plant display name
        log_folder: Folder path for saving logs
        verbose: If True, print diagnostic tables (asset snapshots, demand
                 breakdowns, etc.).  Set False for FY batch mode to skip ~20
                 redundant DB queries that are only for display.
    """
    # ── Fetch and display (wrapped in LogCapture for file logging) ────────
    with LogCapture() as cap:
        print(f"\nFetching data for {_MONTH_NAMES.get(month, month)} {year}...")
        print("#" * 70)
        print(f"  NMD CPP  [{plant_id[:8]}]  Period: {month}/{year}")
        print("#" * 70)

        if verbose:
            _print_plant_info()
            _print_power_assets(plant_id)
            _print_asset_snapshot(plant_id, month, year)
            _print_gt_heat_rate_and_free_steam(plant_id, month, year)
            _print_steam_assets(plant_id, month, year)
            _print_consolidated_demand(plant_id, month, year)
            _print_import_power(plant_id, month, year)
            _print_supply_vs_demand(plant_id, month, year)
            _print_generation_utilities_and_u4u(month, year)
            _print_cpp_norms_for_u4u(month, year)

        # ── Run U4U Iterative Balancing ────────────────────────────────────
        print("\n" + "=" * 70)
        print("  U4U ITERATIVE BALANCING")
        print("=" * 70)

        NMDNormsReader.clear_cache()
        result = run_u4u_iteration(plant_id, month, year)

        print(f"\n  Iterations: {result['iterations_used']}  |  "
              f"Converged: {'YES' if result['converged'] else 'NO'}")
        if result['pass_through_utilities']:
            print(f"  Pass-through utilities: {', '.join(result['pass_through_utilities'])}")

        print("\n" + "=" * 70)
        print("  DONE")
        print("=" * 70)

        # ── U4U vs BPC Comparison (with Excel generation) ───────────────────────
        if result:
            _print_u4u_bpc_comparison(result, month, year)

        # ── Norms Save ────────────────────────────────────────────────────
        if result:
            save_calculated_norms(month, year, result, dry_run=dry_run)

        # ── Cost Module — Calculate utility prices ─────────────────────────────
        fy_start = year if month >= 4 else year - 1
        fy_str = f"{fy_start}-{str(fy_start + 1)[-2:]}"
        price_result = calculate_and_print_utility_prices(month, year,
                                           cpp_plant_id=plant_id,
                                           financial_year=fy_str,
                                           dry_run=dry_run)

    # ── Save log file ─────────────────────────────────────────────────────
    log_path = None
    try:
        log_path = save_text_log(cap.text, plant_name, month, year, log_folder)
        print(f"\nLog saved: {log_path}")
    except Exception as e:
        print(f"\n[WARNING] Could not save log: {e}")

    return {
        "month": month,
        "year": year,
        "u4u_result": result,
        "price_result": price_result,
        "log_path": log_path,
        "log_text": cap.text,
    }


# ---------------------------------------------------------------------------
# Cost-only helpers
# ---------------------------------------------------------------------------

def _run_cost_only_month(month, year, plant_id, plant_name, log_folder, dry_run=False):
    """Run only the cost module for a single month (no balance model)."""
    fy_start = year if month >= 4 else year - 1
    fy_str = f"{fy_start}-{str(fy_start + 1)[-2:]}"

    with LogCapture() as cap:
        price_result = calculate_and_print_utility_prices(
            month, year, cpp_plant_id=plant_id, financial_year=fy_str, dry_run=dry_run
        )

    log_path = None
    try:
        log_path = save_text_log(cap.text, plant_name, month, year, log_folder)
        print(f"\n  Log saved: {log_path}")
    except Exception as e:
        print(f"\n  [WARNING] Could not save log: {e}")

    return {
        "month": month,
        "year": year,
        "u4u_result": None,
        "price_result": price_result,
        "log_path": log_path,
        "log_text": cap.text,
    }


def _generate_yearly_bpc_comparison(month_results, fy_log_folder, fy_label):
    """Generate yearly BPC price comparison log from cost-only results."""
    price_comparison_log_path = None
    try:
        month_tables = []
        yearly_rows = []

        for mr in month_results:
            m = mr["month"]
            y = mr["year"]
            m_name = _MONTH_NAMES.get(m, m)
            pr = mr.get("price_result") or {}

            if not pr or not pr.get("success"):
                month_tables.append(
                    "\n".join([
                        "=" * 196,
                        f"  CPP vs BPC — UTILITY PRICE COMPARISON ({m_name} {y})",
                        "=" * 196,
                        "  [BPC COMPARISON] Skipped — utility price result not available",
                    ])
                )
                continue

            bpc_reader = BPCODSReader.get_reader(m, y)
            bpc_reader.load()
            if not bpc_reader.is_available:
                month_tables.append(
                    "\n".join([
                        "=" * 196,
                        f"  CPP vs BPC — UTILITY PRICE COMPARISON ({m_name} {y})",
                        "=" * 196,
                        "  [BPC COMPARISON] Skipped — BPC.ods not available",
                    ])
                )
                continue

            table_text, rows, _ = build_bpc_comparison_table_text(
                month=m,
                year=y,
                nmd_groups=pr.get("nmd_groups", {}),
                utility_prices=pr.get("utility_prices", {}),
                calc_sequence=pr.get("calc_sequence", []),
                bpc_reader=bpc_reader,
                title=f"CPP vs BPC — UTILITY PRICE COMPARISON ({m_name} {y})",
            )
            if table_text:
                month_tables.append(table_text)
                yearly_rows.extend(rows)
            else:
                month_tables.append(
                    "\n".join([
                        "=" * 196,
                        f"  CPP vs BPC — UTILITY PRICE COMPARISON ({m_name} {y})",
                        "=" * 196,
                        "  [BPC COMPARISON] Skipped — NMD data not available",
                    ])
                )

        if yearly_rows:
            yearly_table = build_yearly_bpc_comparison_table_text(
                yearly_rows,
                title=f"CPP vs BPC — UTILITY PRICE COMPARISON ({fy_label} TOTAL)",
            )
            price_comparison_log_path = os.path.join(fy_log_folder, "utility_price_bpc_comparison.log")
            with open(price_comparison_log_path, "w", encoding="utf-8") as f:
                f.write(f"UTILITY PRICE COMPARISON LOG - {fy_label}\n")
                f.write("=" * 196 + "\n\n")
                for table in month_tables:
                    f.write(table + "\n\n")
                f.write(yearly_table + "\n")
            print(f"\n  Utility price comparison log saved: {price_comparison_log_path}")
    except Exception as exc:
        print(f"\n  [WARN] Utility price comparison log generation failed: {exc}")

    # Summary
    summary_path = os.path.join(fy_log_folder, "summary.txt")
    with open(summary_path, "w", encoding="utf-8") as f:
        f.write(f"COST-ONLY RUN SUMMARY - {fy_label}\n")
        f.write("=" * 120 + "\n\n")
        f.write(f"CPP Plant: {plant_name}\n")
        f.write(f"Months processed: {len(month_results)}\n\n")

        f.write("MONTH-BY-MONTH QUICK RESULTS:\n")
        f.write("-" * 120 + "\n")
        for mr in month_results:
            m = mr["month"]
            y = mr["year"]
            m_name = _MONTH_NAMES.get(m, m)
            pr = mr.get("price_result") or {}
            price_ok = pr.get("success", False)
            price_iters = pr.get("iterations", 0)
            converged = pr.get("converged", False)

            status = "[OK]" if price_ok else "[FAIL]"
            f.write(f"{status} {m_name} {y}: "
                    f"Price={'YES' if price_ok else 'NO'} (iters={price_iters}, "
                    f"converged={'YES' if converged else 'NO'}) | "
                    f"Log: {mr.get('log_path', 'N/A')}\n")

        f.write("\n")
        if price_comparison_log_path:
            f.write(f"Utility price comparison log: {price_comparison_log_path}\n")

    print(f"  Summary saved: {summary_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    setup_logging()

    parser = argparse.ArgumentParser(
        description="NMD CPP — Fetch & Display Asset Details",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    python main.py
    python main.py --month 4 --year 2026
    python main.py --fy 2025  (Execute full financial year April 2025 to March 2026)
    python main.py --cost-only --month 4 --year 2026  (Cost module only, single month)
    python main.py --cost-only --fy 2025  (Cost module only, full financial year)
    python main.py --dry-run --month 4 --year 2025  (Dry run, no DB writes)
    python main.py --dry-run --fy 2025  (Dry run, full financial year)
        """,
    )
    parser.add_argument("--month", type=int, default=None, help="Month (1-12) - for single month execution")
    parser.add_argument("--year",  type=int, default=None, help="Calendar year - for single month execution")
    parser.add_argument("--fy", type=int, default=None, help="Financial year (YYYY) - for full year execution (April to March)")
    parser.add_argument("--cost-only", action="store_true", help="Run only the cost module (utility price calculation + BPC comparison) without the balance model. Requires norms already saved for the target month(s).")
    parser.add_argument("--dry-run", action="store_true", help="Skip all DB writes (norms save, price save, snapshot). Use to preview calculations without modifying any tables.")
    args = parser.parse_args()

    plant_id = NMD_PLANT_ID
    plant_name = PLANT_REGISTRY[plant_id]["display_name"]

    print("=" * 70)
    print("NMD CPP — FETCH ASSET DETAILS")
    print("=" * 70)
    print(f"  Plant: {plant_name}")

    # ── Resolve execution mode: single month vs financial year ─────────────
    fy = args.fy
    month = args.month
    year = args.year

    cost_only = args.cost_only
    dry_run = args.dry_run

    if dry_run:
        print(f"  ** DRY RUN MODE — no DB writes **")

    if cost_only:
        # ── Cost-only mode: skip balance model, run just cost module ───────
        print(f"\n  Mode: COST ONLY (no balance model)")

        if fy:
            months_to_execute = []
            for m in range(4, 13):
                months_to_execute.append((m, fy))
            for m in range(1, 4):
                months_to_execute.append((m, fy + 1))
            run_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            fy_base_folder = os.path.join(LOG_FOLDER, f"{plant_name.replace(' ', '_')}_FY{fy}-{fy+1}")
            fy_log_folder = os.path.join(fy_base_folder, f"cost_only_{run_timestamp}")
            os.makedirs(fy_log_folder, exist_ok=True)
            fy_label = f"FY {fy}-{str(fy + 1)[-2:]}"
            print(f"  Financial Year: {fy_label}")
            print(f"  Log folder: {fy_log_folder}")
            print("=" * 70)

            month_results = []
            for idx, (exec_month, exec_year) in enumerate(months_to_execute, 1):
                print(f"\n{'='*70}")
                print(f"  MONTH {idx}/{len(months_to_execute)}: {_MONTH_NAMES.get(exec_month, exec_month)} {exec_year}")
                print(f"{'='*70}")
                mr = _run_cost_only_month(exec_month, exec_year, plant_id, plant_name, fy_log_folder, dry_run=dry_run)
                month_results.append(mr)

            # Generate yearly BPC comparison log
            _generate_yearly_bpc_comparison(month_results, fy_log_folder, fy_label)

            print(f"\n{'='*70}")
            print(f"  COST-ONLY RUN COMPLETED — {fy_label}")
            print(f"  Logs: {fy_log_folder}")
            print(f"{'='*70}")

        else:
            if not month or not year:
                print("\n--- PERIOD ---")
                if not month:
                    month_input = input("Enter Month (1-12): ").strip()
                    month = int(month_input) if month_input else None
                if not year:
                    year_input = input("Enter Year: ").strip()
                    year = int(year_input) if year_input else None
            if not month or not year:
                print("\nERROR: Must provide both month and year")
                parser.print_help()
                sys.exit(1)

            plant_log = os.path.join(LOG_FOLDER, plant_name.replace(" ", "_"), "cost_only")
            os.makedirs(plant_log, exist_ok=True)
            _run_cost_only_month(month, year, plant_id, plant_name, plant_log, dry_run=dry_run)

    elif fy:
        # Financial year mode
        print(f"\n  Mode: FINANCIAL YEAR EXECUTION")
        print(f"  Financial Year: {fy}-{fy+1}")
        print(f"  Period: April {fy} to March {fy+1}")
        # Financial year months: April (4) to March (3) of next year
        months_to_execute = []
        for m in range(4, 13):  # April to December
            months_to_execute.append((m, fy))
        for m in range(1, 4):   # January to March of next year
            months_to_execute.append((m, fy + 1))

        # Create timestamped run folder inside the FY folder
        run_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        fy_base_folder = os.path.join(LOG_FOLDER, f"{plant_name.replace(' ', '_')}_FY{fy}-{fy+1}")
        fy_log_folder = os.path.join(fy_base_folder, f"run_{run_timestamp}")
        os.makedirs(fy_log_folder, exist_ok=True)

        fy_label = f"FY {fy}-{str(fy + 1)[-2:]}"

        print(f"\n  Executing {len(months_to_execute)} months sequentially...")
        print(f"  Log folder: {fy_log_folder}")
        print("=" * 70)

        # Collect results for yearly comparison
        month_results = []

        # Execute each month
        for idx, (exec_month, exec_year) in enumerate(months_to_execute, 1):
            print(f"\n{'='*70}")
            print(f"  MONTH {idx}/{len(months_to_execute)}: {_MONTH_NAMES.get(exec_month, exec_month)} {exec_year}")
            print(f"{'='*70}")

            # Execute single month (verbose=False: skip diagnostic prints in FY mode)
            month_result = execute_single_month(exec_month, exec_year, plant_id, plant_name, fy_log_folder,
                                 verbose=False, dry_run=dry_run)
            month_results.append(month_result)

            print(f"\n  ✓ Completed {_MONTH_NAMES.get(exec_month, exec_month)} {exec_year}")

        # ── Generate yearly BPC price comparison log ───────────────────────
        price_comparison_log_path = None
        try:
            month_tables = []
            yearly_rows = []

            for mr in month_results:
                m = mr["month"]
                y = mr["year"]
                m_name = _MONTH_NAMES.get(m, m)
                pr = mr.get("price_result") or {}

                if not pr or not pr.get("success"):
                    month_tables.append(
                        "\n".join([
                            "=" * 196,
                            f"  CPP vs BPC — UTILITY PRICE COMPARISON ({m_name} {y})",
                            "=" * 196,
                            "  [BPC COMPARISON] Skipped — utility price result not available",
                        ])
                    )
                    continue

                bpc_reader = BPCODSReader.get_reader(m, y)
                bpc_reader.load()
                if not bpc_reader.is_available:
                    month_tables.append(
                        "\n".join([
                            "=" * 196,
                            f"  CPP vs BPC — UTILITY PRICE COMPARISON ({m_name} {y})",
                            "=" * 196,
                            "  [BPC COMPARISON] Skipped — BPC.ods not available",
                        ])
                    )
                    continue

                table_text, rows, _ = build_bpc_comparison_table_text(
                    month=m,
                    year=y,
                    nmd_groups=pr.get("nmd_groups", {}),
                    utility_prices=pr.get("utility_prices", {}),
                    calc_sequence=pr.get("calc_sequence", []),
                    bpc_reader=bpc_reader,
                    title=f"CPP vs BPC — UTILITY PRICE COMPARISON ({m_name} {y})",
                )
                if table_text:
                    month_tables.append(table_text)
                    yearly_rows.extend(rows)
                else:
                    month_tables.append(
                        "\n".join([
                            "=" * 196,
                            f"  CPP vs BPC — UTILITY PRICE COMPARISON ({m_name} {y})",
                            "=" * 196,
                            "  [BPC COMPARISON] Skipped — NMD data not available",
                        ])
                    )

            if yearly_rows:
                yearly_table = build_yearly_bpc_comparison_table_text(
                    yearly_rows,
                    title=f"CPP vs BPC — UTILITY PRICE COMPARISON ({fy_label} TOTAL)",
                )
                price_comparison_log_path = os.path.join(fy_log_folder, "utility_price_bpc_comparison.log")
                with open(price_comparison_log_path, "w", encoding="utf-8") as f:
                    f.write(f"UTILITY PRICE COMPARISON LOG - {fy_label}\n")
                    f.write("=" * 196 + "\n\n")
                    for table in month_tables:
                        f.write(table + "\n\n")
                    f.write(yearly_table + "\n")
                print(f"\n  Utility price comparison log saved: {price_comparison_log_path}")
        except Exception as exc:
            print(f"\n  [WARN] Utility price comparison log generation failed: {exc}")

        # ── Generate summary.txt ───────────────────────────────────────────
        summary_path = os.path.join(fy_log_folder, "summary.txt")
        with open(summary_path, "w", encoding="utf-8") as f:
            f.write(f"FULL FINANCIAL YEAR RUN SUMMARY - {fy_label}\n")
            f.write("=" * 120 + "\n\n")
            f.write(f"CPP Plant: {plant_name}  [{plant_id[:8]}]\n")
            f.write(f"Run timestamp: {run_timestamp}\n")
            f.write(f"Months processed: {len(months_to_execute)}\n\n")

            f.write("MONTH-BY-MONTH QUICK RESULTS:\n")
            f.write("-" * 120 + "\n")
            for mr in month_results:
                m = mr["month"]
                y = mr["year"]
                m_name = _MONTH_NAMES.get(m, m)
                pr = mr.get("price_result") or {}
                u4u = mr.get("u4u_result") or {}
                converged = u4u.get("converged", False)
                iters = u4u.get("iterations_used", 0)
                price_ok = pr.get("success", False)
                price_iters = pr.get("iterations", 0)

                status = "[OK]" if (converged and price_ok) else "[WARN]"
                f.write(f"{status} {m_name} {y}: "
                        f"U4U={'YES' if converged else 'NO'} (iters={iters}) | "
                        f"Price={'YES' if price_ok else 'NO'} (iters={price_iters}) | "
                        f"Log: {mr.get('log_path', 'N/A')}\n")

            f.write("\n")
            if price_comparison_log_path:
                f.write(f"Utility price comparison log: {price_comparison_log_path}\n")

        print(f"\n  Summary saved: {summary_path}")

        print(f"\n{'='*70}")
        print(f"  FINANCIAL YEAR {fy}-{fy+1} EXECUTION COMPLETED")
        print(f"  Total months processed: {len(months_to_execute)}")
        print(f"  Logs saved in: {fy_log_folder}")
        if price_comparison_log_path:
            print(f"  Price comparison log: {price_comparison_log_path}")
        print(f"  Summary: {summary_path}")
        print(f"{'='*70}")
        
    else:
        # Single month mode (existing logic)
        if not month or not year:
            print("\n--- PERIOD ---")
            if not month:
                month_input = input("Enter Month (1-12): ").strip()
                month = int(month_input) if month_input else None
            if not year:
                year_input = input("Enter Year: ").strip()
                year = int(year_input) if year_input else None

        if not month or not year:
            print("\nERROR: Must provide both month and year")
            parser.print_help()
            sys.exit(1)
        
        # Create standard log folder
        plant_log = os.path.join(LOG_FOLDER, plant_name.replace(" ", "_"))
        os.makedirs(plant_log, exist_ok=True)
        
        execute_single_month(month, year, plant_id, plant_name, plant_log, dry_run=dry_run)


if __name__ == "__main__":
    main()
