"""
Shared Excel report generator for all JMD plants.

Generates two report types from a calculation result:
  - Monthly report  : one sheet per section for a single month
  - Full-year report: one sheet per month, plus a FY summary sheet

Usage:
    from engine.excel_report import write_month_report, write_full_year_report
"""

import os
import logging
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------

_MONTHS = ["Apr", "May", "Jun", "Jul", "Aug", "Sep",
           "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]

_MONTH_NUM = {
    1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr",
    5: "May", 6: "Jun", 7: "Jul", 8: "Aug",
    9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec",
}

if _HAS_OPENPYXL:
    _HDR_FILL    = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    _SUB_FILL    = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    _TOTAL_FILL  = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    _HDR_FONT    = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    _BOLD_FONT   = Font(name="Calibri", size=11, bold=True)
    _NORMAL_FONT = Font(name="Calibri", size=10)
    _THIN = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )


# ---------------------------------------------------------------------------
# Low-level helpers
# ---------------------------------------------------------------------------

def _style_cell(cell, value=None, font=None, fill=None, align="left", border=True):
    if not _HAS_OPENPYXL:
        return
    if value is not None:
        cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if border:
        cell.border = _THIN


def _header_row(ws, col_start, row, values: list):
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=col_start + i)
        _style_cell(c, v, font=_HDR_FONT, fill=_HDR_FILL, align="center")


def _data_row(ws, row, values: list, bold=False, fill=None):
    font = _BOLD_FONT if bold else _NORMAL_FONT
    for col, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=col)
        align = "right" if isinstance(v, (int, float)) else "left"
        _style_cell(c, v, font=font, fill=fill, align=align)


def _section_title(ws, row, title: str, ncols: int):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1)
    _style_cell(c, title, font=_HDR_FONT, fill=_HDR_FILL, align="left")


def _auto_width(ws, min_w=8, max_w=40):
    for col_cells in ws.columns:
        length = max(
            len(str(c.value or "")) for c in col_cells
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = \
            min(max(length + 2, min_w), max_w)


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------

def _write_plant_info(ws, result: dict):
    ws.title = "Plant Info"
    info_rows = [
        ("Plant ID",    result.get("plant_id", "")),
        ("Plant Name",  result.get("plant_name", "")),
        ("FY",          result.get("fy", "")),
        ("Month",       result.get("month", "")),
        ("Year",        result.get("year", "")),
        ("Generated",   datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    _header_row(ws, 1, 1, ["Field", "Value"])
    for r, (k, v) in enumerate(info_rows, start=2):
        _data_row(ws, r, [k, str(v)])
    _auto_width(ws)


def _write_operational_hours(ws, result: dict):
    ws.title = "Operational Hours"
    rows = result.get("op_hours", [])
    if not rows:
        ws.cell(row=1, column=1).value = "No data"
        return

    headers = ["Asset", "Type", "AOP Year"] + _MONTHS + ["Total"]
    _header_row(ws, 1, 1, headers)

    for r, row in enumerate(rows, start=2):
        vals = [row["asset_name"], row.get("asset_type", ""), row.get("aop_year", "")]
        month_vals = [row.get(m) for m in _MONTHS]
        total = sum(v for v in month_vals if v is not None)
        formatted = [round(v, 1) if v is not None else "" for v in month_vals]
        fill = _TOTAL_FILL if r % 2 == 0 else None
        _data_row(ws, r, vals + formatted + [round(total, 1)], fill=fill)

    _auto_width(ws)


def _write_power_priority(ws, result: dict):
    ws.title = "Power Priority"
    rows = result.get("power_priority", [])
    if not rows:
        ws.cell(row=1, column=1).value = "No data"
        return

    headers = ["Asset", "Type", "AOP Year"] + _MONTHS
    _header_row(ws, 1, 1, headers)

    for r, row in enumerate(rows, start=2):
        vals = [row["asset_name"], row.get("asset_type", ""), row.get("aop_year", "")]
        pri_vals = [row.get(m, "") for m in _MONTHS]
        fill = _SUB_FILL if r % 2 == 0 else None
        _data_row(ws, r, vals + pri_vals, fill=fill)

    _auto_width(ws)


def _write_steam_priority(ws, result: dict):
    ws.title = "Steam Priority"
    rows = result.get("steam_priority", [])
    if not rows:
        ws.cell(row=1, column=1).value = "No data"
        return

    headers = ["Asset", "Type", "Steam Type", "AOP Year"] + _MONTHS
    _header_row(ws, 1, 1, headers)

    for r, row in enumerate(rows, start=2):
        vals = [row["asset_name"], row.get("asset_type", ""),
                row.get("steam_type", ""), row.get("aop_year", "")]
        pri_vals = [row.get(m, "") for m in _MONTHS]
        fill = _SUB_FILL if r % 2 == 0 else None
        _data_row(ws, r, vals + pri_vals, fill=fill)

    _auto_width(ws)


def _write_demands(ws, result: dict):
    ws.title = "Demands"
    demands = result.get("demands", {})
    if not demands:
        ws.cell(row=1, column=1).value = "No data"
        return

    _header_row(ws, 1, 1, ["Parameter", "Value"])
    for r, (k, v) in enumerate(demands.items(), start=2):
        _data_row(ws, r, [k, round(v, 4) if isinstance(v, float) else v])

    _auto_width(ws)


def _write_import_power(ws, result: dict):
    ws.title = "Import Power"
    imp = result.get("import_power", {})
    if not imp:
        ws.cell(row=1, column=1).value = "No data"
        return

    _header_row(ws, 1, 1, ["Source", "Capacity MW", "Hours", "MWh"])
    per_source = imp.get("per_source", [])
    for r, s in enumerate(per_source, start=2):
        _data_row(ws, r, [
            s.get("source_name", ""),
            s.get("capacity_mw", 0),
            s.get("hours", 0),
            s.get("mwh", 0),
        ])

    total_row = len(per_source) + 2
    _data_row(ws, total_row,
              ["TOTAL", "", "", imp.get("total_mwh", 0)],
              bold=True, fill=_TOTAL_FILL)
    _auto_width(ws)


def _write_norms(ws, result: dict):
    ws.title = "Norms"
    norms = result.get("norms", [])
    if not norms:
        ws.cell(row=1, column=1).value = "No data"
        return

    # Dynamic columns from first row
    cols = list(norms[0].keys()) if norms else []
    _header_row(ws, 1, 1, cols)
    for r, row in enumerate(norms, start=2):
        vals = [row.get(c, "") for c in cols]
        _data_row(ws, r, vals)

    _auto_width(ws)


def _write_assets(ws, result: dict):
    ws.title = "Assets"
    assets = result.get("assets", [])
    if not assets:
        ws.cell(row=1, column=1).value = "No data"
        return

    headers = ["Asset ID", "Asset Name", "Display Name", "Type",
               "Plant Code", "Utility Gen FK", "Utility Dist FK"]
    _header_row(ws, 1, 1, headers)
    for r, a in enumerate(assets, start=2):
        _data_row(ws, r, [
            a.get("asset_id", ""), a.get("asset_name", ""),
            a.get("display_name", ""), a.get("asset_type", ""),
            a.get("plant_code", ""),
            a.get("utility_generation_fk", ""), a.get("utility_distributed_fk", ""),
        ])
    _auto_width(ws)


def _write_u4u_consumption(ws, result: dict):
    """Write detailed U4U consumption table — one row per producer+material."""
    ws.title = "U4U Consumption"
    u4u = result.get("u4u_iteration") or {}
    records = u4u.get("final_detail_records", [])
    if not records:
        ws.cell(row=1, column=1).value = "No U4U data"
        return

    headers = [
        "Utility Plant", "Utility", "UOM", "Gen Qty", "BPC Gen Qty", "Gen Diff %",
        "Account", "Material", "Material UOM", "Norm",
        "Quantity", "BPC Quantity", "Qty Diff %",
    ]
    _header_row(ws, 1, 1, headers)

    # Build BPC lookup from u4u result
    bpc_gen = u4u.get("final_bpc_gen_quantities", {})
    bpc_qty = u4u.get("final_bpc_quantities", {})

    for r, rec in enumerate(records, start=2):
        p = rec["producer"]
        bg = bpc_gen.get(p, 0.0)
        bq = (bpc_qty.get(p) or {}).get(rec["material"], 0.0)
        gen_diff = _pct_diff(rec["generation"], bg)
        qty_diff = _pct_diff(rec["quantity"], bq)
        _data_row(ws, r, [
            p,
            rec.get("producer_utility", ""),
            rec.get("producer_uom", ""),
            round(rec["generation"], 2),
            round(bg, 2),
            round(gen_diff, 2),
            rec.get("account", ""),
            rec.get("material", ""),
            rec.get("material_uom", ""),
            round(rec.get("norm", 0.0), 6),
            round(rec["quantity"], 2),
            round(bq, 2),
            round(qty_diff, 2),
        ])

    # Freeze header + auto filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(records) + 1}"
    _auto_width(ws)


def _write_u4u_summary(ws, result: dict):
    """Write per-utility summary comparison — one row per generation utility."""
    ws.title = "U4U Summary"
    u4u = result.get("u4u_iteration") or {}
    records = u4u.get("final_detail_records", [])
    if not records:
        ws.cell(row=1, column=1).value = "No U4U data"
        return

    bpc_gen = u4u.get("final_bpc_gen_quantities", {})

    # Aggregate per producer
    producer_summary: dict = {}
    for rec in records:
        p = rec["producer"]
        if p not in producer_summary:
            producer_summary[p] = {
                "utility": rec.get("producer_utility", ""),
                "uom": rec.get("producer_uom", ""),
                "generation": rec["generation"],
                "bpc_gen": bpc_gen.get(p, 0.0),
                "material_count": 0,
                "materials": [],
            }
        producer_summary[p]["material_count"] += 1
        producer_summary[p]["materials"].append(rec.get("material", ""))

    headers = [
        "Utility Plant", "Utility", "UOM", "Gen Qty", "BPC Gen Qty",
        "Gen Diff %", "# Materials", "Materials Consumed",
    ]
    _header_row(ws, 1, 1, headers)

    for r, p in enumerate(sorted(producer_summary.keys()), start=2):
        s = producer_summary[p]
        gen_diff = _pct_diff(s["generation"], s["bpc_gen"])
        fill = _SUB_FILL if r % 2 == 0 else None
        _data_row(ws, r, [
            p,
            s["utility"],
            s["uom"],
            round(s["generation"], 2),
            round(s["bpc_gen"], 2),
            round(gen_diff, 2),
            s["material_count"],
            ", ".join(s["materials"]),
        ], fill=fill)

    # Totals row
    total_row = len(producer_summary) + 2
    total_gen = sum(s["generation"] for s in producer_summary.values())
    total_bpc = sum(s["bpc_gen"] for s in producer_summary.values())
    _data_row(ws, total_row, [
        "TOTAL", "", "", round(total_gen, 2), round(total_bpc, 2),
        "", sum(s["material_count"] for s in producer_summary.values()),
        "",
    ], bold=True, fill=_TOTAL_FILL)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{total_row}"
    _auto_width(ws)


def _pct_diff(actual: float, bpc: float) -> float:
    """Calculate percentage difference between actual and BPC value."""
    if bpc == 0:
        return 0.0
    return (actual - bpc) / bpc * 100.0


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def write_month_report(result: dict, output_folder: str) -> str:
    """
    Write a single-month Excel report.

    Args:
        result:        dict returned by engine.calculator.run_month()
        output_folder: directory to save the file

    Returns:
        Absolute path to the generated .xlsx file.
    """
    if not _HAS_OPENPYXL:
        raise ImportError("openpyxl is required: pip install openpyxl")

    plant_name = result.get("plant_name", "plant").replace(" ", "_")
    month      = result.get("month", 0)
    year       = result.get("year", 0)
    fy         = result.get("fy", "")
    ts         = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename   = f"{plant_name}_{year}_{month:02d}_{ts}.xlsx"

    os.makedirs(output_folder, exist_ok=True)
    filepath = os.path.join(output_folder, filename)

    wb = Workbook()
    wb.remove(wb.active)

    _write_plant_info(wb.create_sheet(), result)
    _write_assets(wb.create_sheet(), result)
    _write_operational_hours(wb.create_sheet(), result)
    _write_power_priority(wb.create_sheet(), result)
    _write_steam_priority(wb.create_sheet(), result)
    _write_demands(wb.create_sheet(), result)
    _write_import_power(wb.create_sheet(), result)
    _write_norms(wb.create_sheet(), result)
    _write_u4u_summary(wb.create_sheet(), result)
    _write_u4u_consumption(wb.create_sheet(), result)

    wb.save(filepath)
    logger.info("  [EXCEL] Month report saved: %s", filepath)
    return filepath


def write_full_year_report(full_year_result: dict, output_folder: str) -> str:
    """
    Write a full-year Excel report with one sheet per month + FY Summary sheet.

    Args:
        full_year_result: dict returned by engine.calculator.run_full_year()
        output_folder:    directory to save the file

    Returns:
        Absolute path to the generated .xlsx file.
    """
    if not _HAS_OPENPYXL:
        raise ImportError("openpyxl is required: pip install openpyxl")

    plant_name = full_year_result.get("plant_name", "plant").replace(" ", "_")
    fy         = full_year_result.get("fy", "")
    ts         = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename   = f"{plant_name}_FY{fy}_{ts}.xlsx"

    os.makedirs(output_folder, exist_ok=True)
    filepath = os.path.join(output_folder, filename)

    wb = Workbook()
    wb.remove(wb.active)

    # ── FY Summary sheet ─────────────────────────────────────────────────────
    summary_ws = wb.create_sheet("FY Summary")
    _header_row(summary_ws, 1, 1,
                ["Month", "Year", "Success", "Assets", "Norms",
                 "Exec Time (s)", "Message"])
    months_dict = full_year_result.get("months", {})
    # Sort Apr → Mar
    ordered_keys = sorted(
        months_dict.keys(),
        key=lambda k: (int(k[:4]), int(k[5:])) if int(k[5:]) >= 4
                      else (int(k[:4]) + 1, int(k[5:]))
    )
    for r, key in enumerate(ordered_keys, start=2):
        res = months_dict[key]
        _data_row(summary_ws, r, [
            _MONTH_NUM.get(res.get("month", 0), ""),
            res.get("year", ""),
            "OK" if res.get("success") else "FAIL",
            len(res.get("assets", [])),
            len(res.get("norms",  [])),
            res.get("execution_time_seconds", ""),
            res.get("message", ""),
        ])
    _auto_width(summary_ws)

    # ── Operational Hours across all months ───────────────────────────────────
    oh_ws = wb.create_sheet("Operational Hours FY")
    # Use first successful month's op_hours (all 12-month rows)
    first_result = next(
        (r for r in months_dict.values() if r.get("success") and r.get("op_hours")),
        None,
    )
    if first_result:
        _write_operational_hours(oh_ws, first_result)
        oh_ws.title = "Operational Hours FY"

    # ── Power Priority across all months ────────────────────────────────────
    pp_ws = wb.create_sheet("Power Priority FY")
    if first_result:
        _write_power_priority(pp_ws, first_result)
        pp_ws.title = "Power Priority FY"

    # ── Steam Priority across all months ────────────────────────────────────
    sp_ws = wb.create_sheet("Steam Priority FY")
    if first_result:
        _write_steam_priority(sp_ws, first_result)
        sp_ws.title = "Steam Priority FY"

    # ── One sheet per month (demands + import power + norms) ─────────────────
    for key in ordered_keys:
        res = months_dict[key]
        month_label = f"{_MONTH_NUM.get(res.get('month', 0), key)} {res.get('year', '')}"
        ws = wb.create_sheet(month_label[:31])  # Excel sheet name limit = 31 chars

        row = 1
        _section_title(ws, row, f"{month_label} — Demands", 8)
        row += 1
        demands = res.get("demands", {})
        _header_row(ws, 1, row, ["Parameter", "Value"])
        row += 1
        for k, v in demands.items():
            _data_row(ws, row, [k, round(v, 4) if isinstance(v, float) else v])
            row += 1

        row += 1
        _section_title(ws, row, f"{month_label} — Import Power", 8)
        row += 1
        imp = res.get("import_power", {})
        _header_row(ws, 1, row, ["Source", "Capacity MW", "Hours", "MWh"])
        row += 1
        for s in imp.get("per_source", []):
            _data_row(ws, row, [
                s.get("source_name", ""), s.get("capacity_mw", 0),
                s.get("hours", 0), s.get("mwh", 0),
            ])
            row += 1
        _data_row(ws, row, ["TOTAL", "", "", imp.get("total_mwh", 0)],
                  bold=True, fill=_TOTAL_FILL)
        row += 2

        _section_title(ws, row, f"{month_label} — Norms", 8)
        row += 1
        norms = res.get("norms", [])
        if norms:
            cols = list(norms[0].keys())
            _header_row(ws, 1, row, cols)
            row += 1
            for norm in norms:
                _data_row(ws, row, [norm.get(c, "") for c in cols])
                row += 1

        _auto_width(ws)

    wb.save(filepath)
    logger.info("  [EXCEL] Full year report saved: %s", filepath)
    return filepath
